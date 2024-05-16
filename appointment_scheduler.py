import csv
import datetime as dt
import json
import math
import numpy as np
import openpyxl 
import psutil
# import xlsxwriter
import requests
import sys
from datetime import datetime
from dateutil import parser as date_parser
from matplotlib.lines import Line2D
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import load_workbook
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal,QSize,QUrl,QThread
from PyQt5.QtGui import QIcon,QFont,QDesktopServices
from PyQt5.QtGui import QColor,QTextCharFormat, QColor
from PyQt5.QtWidgets \
import QApplication,QCheckBox,QWidgetAction,QDateEdit,QMessageBox,QSpinBox,\
QTableWidget,QLineEdit, QTableWidgetItem, QMainWindow, QHBoxLayout, QVBoxLayout,\
QCalendarWidget, QWidget, QPushButton, QLabel, QInputDialog,  QVBoxLayout,QFileDialog, \
QFormLayout, QScrollArea,QLabel, QSizePolicy,QHeaderView,QGroupBox ,QComboBox, QTextEdit, QFrame,  QDialog, \
QComboBox, QTabWidget, QMenu, QAction, QVBoxLayout,QAbstractItemView

#Globals
book_from_complaint = False 
current_slot =      { 
    
            "date": "",
            "duration": "",  
            "engineer": "",
            "location": ""
                }
class TimeSlotDialog(QDialog):
        '''
        Dialog window for selecting time slots.
        '''
        def __init__(self, locations, engineers, slot={}, parent=None):
            """
            Intialize The BookSlot Dialog Window
            Args:
                locations (list): A list of available locations.
                engineers (list): A list of available engineers.
                parent (QWidget): The parent widget of the dialog window. Default is None.
            Raises:
                None
            Returns:
                None
            """
            super().__init__(parent)
            global current_slot
            self.setWindowTitle("Book Slot")
            self.setGeometry(200, 200, 400, 200)
            self.parent = parent
            self.location_dropdown = QComboBox(self) 
            self.location_dropdown.addItems(locations)
            self.engineer_dropdown = QComboBox(self)
            self.engineer_dropdown.addItems(engineers)
            self.time_options = list(range(1, 9))
            self.time_dropdown = QComboBox(self)
            self.time_dropdown.addItems([f"{hours} Hour(s)" for hours in self.time_options])
            self.time_dropdown.setCurrentIndex(7)
            icon_width = 20
            icon_height = 20
            self.new_button = QPushButton(self)
            self.new_button.setIcon(self.createIcon("file-plus-icon.png", QSize(icon_width, icon_height)))
            self.new_button.setFixedSize(icon_width + 10, icon_height + 10)  
            self.new_button.clicked.connect(self.add_engineer)
            self.new_button1 = QPushButton(self)
            self.new_button1.setIcon(self.createIcon("file-plus-icon.png", QSize(icon_width, icon_height)))
            self.new_button1.setFixedSize(icon_width + 10, icon_height + 10)  
            self.new_button1.clicked.connect(self.add_location)
            self.main_layout = QVBoxLayout(self)
            self.engineer_layout = QHBoxLayout()
            self.engineer_layout.addWidget(self.engineer_dropdown)
            self.engineer_layout.addWidget(self.new_button)
            self.location_layout = QHBoxLayout()
            self.location_layout.addWidget(self.location_dropdown)
            self.location_layout.addWidget(self.new_button1)
            self.main_layout.addLayout(self.engineer_layout)
            self.main_layout.addLayout(self.location_layout)
            self.main_layout.addWidget(self.time_dropdown)
            self.ok_button = QPushButton("OK", self)
            self.ok_button.clicked.connect(self.accept)
            self.main_layout.addWidget(self.ok_button)
            self.cancel_button = QPushButton("Cancel", self)
            self.cancel_button.clicked.connect(self.reject)
            self.main_layout.addWidget(self.cancel_button)
            self.engineers = {} 
            self.locations = {}

            if slot is None:
                slot = {}
            # self.data = self.parent.data
            if isinstance(slot, dict) and len(slot.keys()) >= 0:
                if "engineer" in slot and slot["engineer"] is not None:
                    engg = slot["engineer"]
                    if engg in self.engineers:
                        current_slot = {}
                        self.engineer_dropdown.setCurrentIndex(
                            self.engineer_dropdown.findText(engg))
            else:
                self.engineer_dropdown.addItem("Select Engineer")
                self.location_dropdown.addItem("Select Location")
                self.engineer_dropdown.setCurrentIndex(self.engineer_dropdown.findText("Select Engineer"))
                self.location_dropdown.setCurrentIndex(self.location_dropdown.findText("Select Location"))
        
        def save_json(self):
            """
            Saves the data stored in 'self.data' attribute to a JSON file named 'booked_slot.json'.
            Args:
                self: The instance of the class.
            Raises:
                IOError: If an error occurs while writing to the JSON file.
            Returns:
                None
            """
            with open ("booked_slot.json","w+") as json_file:
                json.dump(self.data,json_file,indent=4)
        
        def load_json(self):
                """
                This method loads data from JSON file named 'booked_slot.json'
                and returns it.
                Args:
                    self: The instance of the class.
                Raises:
                    FileNotFoundError: If the 'booked_slot.json' file not found.
                    JSONDecodeError: If the data in JSON file is not valid JSON.
                Returns:
                    dict: Data loaded from the JSON file.
                """
                with open("booked_slot.json") as json_file:
                    data = json.load(json_file)
                return data       
        
        def createIcon(self, path, size):
            """
            This method creates an icon with specified path and size.
            Args:
                path(str): The file path of the icon.
                size(Qsize): The size of the icon.
            Raises:
                None
            Returns:
                QIcon:An object created from specified path and size.
            """
            icon = QIcon(path)
            pixmap = icon.pixmap(size)
            return QIcon(pixmap)
        
        def add_engineer(self):
            """
            This method allows the user to enter the name of new engineer.If the
            input is valid and not empty the engineer is added to the engineer 
            dropdown menu and the data dict. An API request is sent to add the 
            engineer to the backend server.If the engineer is added sucessfully
            success method is displyed else error message is shown.
            Args:
                self: The instance of the class.
            Raises:
                None
            Returns:
                None
            """
            engineer, ok = QInputDialog.getText(self, "Add Engineer", "Enter the engineer's name:")
            if ok and engineer:
                self.engineer_dropdown.addItem(engineer)
            
                if 'engineers' not in self.data:
                    self.data['engineers'] = {}
                self.data['engineers'][engineer] = []
                api_url = "http://192.168.17.72:5000/api/add_engineer"
                response = requests.post(api_url, json={"engineer": engineer})
                self.save_json()
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Success", "Engineer added successfully!")
                else:
                    QMessageBox.warning(self, "Error", "Engineer added successfully!")

        def add_location(self):
            """
            This method allows the user to enter name of new location.If the
            input is valid and not empty the location is added to location
            dropdown menu and data dict.An API request is sent to add the 
            location to backend server.If the location is added sucessfully
            success method is displayed else error message is shown.
            Args:
                self: The instance of class.
            Raises:
                None
            Returns:
                None
            """
            location, ok = QInputDialog.getText(self, "Add Location", "Enter the location name:")
            if ok and location:
                self.location_dropdown.addItem(location) 

                if 'locations' not in self.data:
                    self.data['locations'] = {}
                self.data['locations'][location] = []
                api_url = "http://192.168.17.72:5000/api/add_location"
                response = requests.post(api_url, json={"location": location})
                self.save_json()
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Success", "Location added successfully!")
                else:
                    QMessageBox.warning(self, "Error", "Location added successfully!")
        
        def exec_(self):
            """
            This method executes the dialog window and returns the result after execution.
            Args:
                self: The instance of class.
            Raises:
                None
            Returns:
                int: The result of executing the dialog window.
            """
            self.result = super().exec_()
            return self.result
        
        def get_selected_data(self):
            """
            This method retrieves the currently selected engineer,location,and time slot
            from the dialog window.
            Args:
                self:The instance of class.
            Raises:
                None
            Returns:
                dict: A dictionary containing selected engineer,location and time slot.
            """
            return {
                'engineer': self.engineer_dropdown.currentText(),
                'location': self.location_dropdown.currentText(), 
                'time': self.time_options[self.time_dropdown.currentIndex()],
            }
        
class DashboardWidget(QWidget):
    """
    Widgets for displaying a dashboard.
    """
    def __init__(self,locations, engineers, parent=None):
        """
        This method initializes the DashboardWidget with various widgets and 
        functionalities for displaying a dashboard.
        Args:
            parent(Qwidget): The parent widget of a dashboard.Default is None.
        Raises:
            None
        Returns:
            None
        """
        super().__init__(parent)
        with open("stylesheet.qss", "r") as style_file:
            self.setStyleSheet(style_file.read())
            self.bookings = []
        with open("booked_slot.json") as json_file:
            data = json.load(json_file)
            engineers_data = data.get("engineers", {})
            locations_data = data.get("locations", {})
            self.dates = sorted(set(entry["date"] for entry in data.get("bookings", [])))
        self.engineer_dropdown = QComboBox(self)
        self.engineer_dropdown.addItem("All Engineer")
        self.engineer_dropdown.addItems(engineers_data.keys())
        self.engineer_dropdown.currentIndexChanged.connect(self.update_charts)
        self.engineer_dropdown.setStyleSheet("background-color: lightblue;")
        self.location_dropdown = QComboBox(self)
        self.location_dropdown.addItem("All Location")
        self.location_dropdown.addItems(locations_data.keys())
        self.location_dropdown.currentIndexChanged.connect(self.update_charts)
        self.location_dropdown.setStyleSheet("background-color: lightblue;")
        self.update_charts_button = QPushButton("Update Charts", self)
        self.update_charts_button.clicked.connect(self.update_charts)
        self.update_charts_button.setStyleSheet("background-color: lightblue;")
        self.update_charts_button.setFixedHeight(20)
        self.engineer_schedule_chart = ScrollableMatplotlibWidget(self, width=800, height=800)
        self.engineer_schedule_chart.setGeometry(800, 800, 800, 600) 
        self.engineer_schedule_chart.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        container_widget = QWidget(self)
        container_layout = QVBoxLayout(container_widget)
        container_layout.addWidget(self.engineer_schedule_chart)
        scroll_area.setWidget(container_widget)
        self.date = QDateEdit(self)
        self.date.setFixedHeight(20)
        self.date.setCalendarPopup(True)
        self.date.setDisplayFormat("yyyy/MM/dd")
        self.date.setDate(QDate.currentDate())
        self.date.dateChanged.connect(self.update_charts)
        self.date.setStyleSheet("background-color: lightblue;")
        self.top_layout = QHBoxLayout()
        self.top_layout.addWidget(self.engineer_dropdown)
        self.top_layout.addWidget(self.location_dropdown)
        self.top_layout.addWidget(self.date)
        self.top_layout.addWidget(self.update_charts_button)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(scroll_area)
        self.layout.addLayout(self.top_layout)
        
    def update_charts(self):
        """
        This method updates charts displayed on the dashboard based on currently selected
        engineer,location,and date filters.
        Args:
            self: The instance of class.
        Raises:
            None
        Returns:
            None
        """
        with open("booked_slot.json") as json_file:
            data = json.load(json_file)
        bookings = data.get("bookings", [])
        selected_engineer = self.engineer_dropdown.currentText()
        selected_location = self.location_dropdown.currentText()
        selected_date = self.date.date().toString("yyyy/MM/dd")
        engineer_schedule_bookings = [
            entry
            for entry in bookings
            if (
                (entry["engineer"] == selected_engineer or selected_engineer == "All Engineer") and
                (entry["location"] == selected_location or selected_location == "All Location") and
                (entry["date"] == selected_date)
            )
        ]
        unique_engineers = sorted(set(entry["engineer"] for entry in engineer_schedule_bookings))
        unique_locations = sorted(set(entry["location"] for entry in engineer_schedule_bookings))
        engineer_data = {engineer: {"locations": [], "durations": []} for engineer in unique_engineers}
        
        for entry in engineer_schedule_bookings:
            engineer_data[entry['engineer']]["locations"].append(entry["location"])
            engineer_data[entry['engineer']]["durations"].append(entry["duration"])
        self.engineer_schedule_chart.plot_engineer_schedule(
            unique_engineers, engineer_data, "Engineer Schedule"
)

class ScrollableMatplotlibWidget(QScrollArea):
    """
    Scrollable widgets for displaying Matplotlib plots.
    """
    def __init__(self, parent=None, width=15, height=10):
        """
        This represents a QScrollArea widgets that contains a Matplotlib plot.
        It allows for displaying Matplotlibs plots in scrollable area.
        Args:
            parent(Qwidget): The parent widget of the ScrollableMatplotWidget. Default is None.
            width(int): The width of the Matplotlib plot in inches. Default is 15.
            height(int): The height of the Matplotlib plot in inches. Default is 10.
        Raises:
            None
        Returns:
            None
        """
        super(ScrollableMatplotlibWidget, self).__init__(parent)
        self.fig, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.fig)
        self.setWidgetResizable(True)
        container_widget = QWidget(self)
        container_layout = QVBoxLayout(container_widget)
        self.figure = Figure(figsize=(width, height))
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)
        container_layout.addWidget(self.canvas)
        self.setWidget(container_widget)
    
    def clear_plot(self):
        """
        This method clears the current plot displayed on the Matplotlib axes object(self.ax)
        Arg:
            self: Instance of the class.
        Raises:
            None
        Returns:
            None
        """
        if self.ax is not None:
            self.ax.clear()
    
    def plot_engineer_schedule(self, engineers, engineer_data, title):
        """
        This method plots the schedule for each engineer based on provided engineer
        data.
        Args:
            engineer(list): A list of engineers.
            engineer_data(dict): A dictionary containing engineer data with keys as
                                 engineer name and values as dictionaries with keys
                                "locations" and "durations".
            title(str): The title of the plot.
        Raises:
            None
        Returns:
            None
                
        """
        self.clear_plot()
        distinct_colors = [
                'tab:blue', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple', 
                'tab:brown', 'tab:pink', 'tab:gray', 'tab:olive', 'tab:cyan', 
                'deepskyblue', 'salmon', 'gold', 'darkviolet', 'limegreen', 
                'darkorange', 'dodgerblue', 'mediumorchid', 'darkslategrey', 
                'darkturquoise', 'rosybrown', 'mediumpurple', 'olivedrab', 
                'steelblue', 'peru', 'royalblue', 'indianred', 'darkkhaki', 
                'mediumseagreen', 'orangered', 'slateblue', 'seagreen', 
                'sienna', 'crimson', 'darkolivegreen', 'teal', 'slategray', 
                'firebrick', 'cadetblue', 'chocolate', 'forestgreen', 
                'mediumaquamarine', 'darkgoldenrod', 'dimgray', 'saddlebrown', 
                'darkcyan', 'darkmagenta', 'mediumslateblue', 'darkorchid']
        bar_width = 0.2     
        unique_locations = set()
        location_color_mapping = {}
        for i, engineer in enumerate(engineers):
            durations = engineer_data[engineer]["durations"]
            locations = engineer_data[engineer]["locations"]
            for unique_location in set(locations) - unique_locations:
                unique_locations.add(unique_location)
                color = distinct_colors[len(unique_locations) % len(distinct_colors)]
                location_color_mapping[unique_location] = color
            for j, (duration, location) in enumerate(zip(durations, locations)):
                x_position = i + j * (bar_width + 0.0013)
                self.ax.bar(x_position, duration, width=bar_width, color=location_color_mapping[location], label=f'{location}')
        self.ax.set_xticks(np.arange(len(engineers)) + (len(engineers) - 2) * 0.025 * (bar_width + 0.124))
        self.ax.set_xticklabels(engineers, rotation=0, ha="center")
        legend_handles = [Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markersize=15, label=label) for label, color in location_color_mapping.items()]
        self.ax.legend(handles=legend_handles, loc='upper right', title='Location')
        self.ax.set_title(title)
        self.ax.set_xlabel('Engineer')
        self.ax.set_ylabel('Duration')
        self.canvas.draw()

class ReportsWidget(QWidget):
    """
    This class represents a Qwidget used for displaying reports.
    """
    def __init__(self, engineers, locations, parent=None):
        """
        This method contains functionalities to filter,display and export booking
        data as a report.
        Args:
            engineers(list): A list of available engineers.
            locations(list): A list of available locations.
            parent(QWidget): The parent widget of the ReportWidget.Default is None.
        Raises:
            None
        Returns:
            None
        """
        super().__init__(parent)
        self.engineers = engineers
        self.locations = locations
        self.data = []  
        self.layout = QVBoxLayout(self)
        self.filter_groupbox = QGroupBox("Filter Options", self)    
        filter_layout = QHBoxLayout(self.filter_groupbox)
        self.engineer_dropdown = QComboBox(self)
        self.engineer_dropdown.addItems(["All"] + list(self.engineers))
        filter_layout.addWidget(QLabel("Engineer:"))
        filter_layout.addWidget(self.engineer_dropdown)
        self.engineer_dropdown.setStyleSheet("background-color: lightblue;")
        self.location_dropdown = QComboBox(self)
        self.location_dropdown.setEditable(True)  # Enable editing
        self.location_dropdown.setInsertPolicy(QComboBox.NoInsert)  # Prevent insertion of new items
        self.location_dropdown.addItems(["All"] + self.locations)
        filter_layout.addWidget(QLabel("Location:"))
        filter_layout.addWidget(self.location_dropdown)
        self.location_dropdown.setStyleSheet("background-color: lightblue;")
        self.from_date_edit = QDateEdit(self)
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("yyyy/MM/dd")
        self.from_date_edit.setDate(QDate.currentDate()) 
        filter_layout.addWidget(QLabel("From Date:"))
        filter_layout.addWidget(self.from_date_edit)
        self.from_date_edit.setStyleSheet("background-color: lightblue;")
        self.to_date_edit = QDateEdit(self)
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("yyyy/MM/dd")
        self.to_date_edit.setDate(QDate.currentDate())  
        filter_layout.addWidget(QLabel("To Date:"))
        filter_layout.addWidget(self.to_date_edit)
        self.to_date_edit.setStyleSheet("background-color: lightblue;")
        self.filter_button = QPushButton("Filter", self)
        self.filter_button.clicked.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_button)
        self.layout.addWidget(self.filter_groupbox)
        self.filter_button.setStyleSheet("background-color: lightblue;")
        # Styling for Filter Group Box
        self.filter_groupbox.setStyleSheet("QGroupBox { font-size: 16px; \
                                        border: 2px solid black; \
                                        border-radius: 10px; \
                                        padding-top: 20px; \
                                        background-color: #f0f0f0; } \
                                        QGroupBox::title { subcontrol-origin: margin; \
                                        subcontrol-position: top center; \
                                        padding: 0 3px; \
                                        color: black; \
                                        font-weight: bold; }")

        # Table
        self.table = QTableWidget(self)
        self.table.setGeometry(800, 800, 800, 600) 
        self.layout.addWidget(self.table)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        # Download button
        self.download_button = QPushButton("Download", self)
        self.download_button.clicked.connect(self.export_to_csv)
        self.download_button.setStyleSheet("background-color: lightblue;")
        self.layout.addWidget(self.download_button)
        # Set the table properties
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        # Load and display booking data
        self.load_booking_data()
        # Display initial data in the table
        self.display_data(self.data)
    
    def load_booking_data(self):
        """
        This method attempts to load booking data from json file named 'booked_slot.json'.
        If the file is found and the JSON data is successfully decoded,the booking data
        is stored in 'data' attribute of the ReportWidget.
        Args:
            self: The instance of class.
        Raises:
            FileNotFoundError: If the 'booked_slot.json' file is not found.
            json.JSONDecodeError: If the JSON data in the file cannot be decoded.
        Returns:
            None
        """
        try:
            with open('booked_slot.json', 'r') as json_file:
                data = json.load(json_file)
                bookings = data.get('bookings', [])
                self.data = bookings
        except FileNotFoundError:
            print(f"Error: File not found ")
        except json.JSONDecodeError:
            print(f"Error: Unable to decode JSON in file")   
    
    def show_filter_dialog(self):
        """
        This method displays a filter dialog containing options to filter booking data by
        engineer,location,and date range.It connects the 'Apply' button to the 'apply_filters'
        method to apply selected filters.
        Args:
            self: The instance of class.
        Raises:
            None
        Returns:
            None
        """
        filter_options_widget = QWidget(self)
        filter_layout = QFormLayout(filter_options_widget)
        engineer_dropdown = QComboBox(self)
        engineer_dropdown.addItems(self.engineers)
        filter_layout.addRow("Engineer:", engineer_dropdown)
        location_dropdown = QComboBox(self)
        location_dropdown.addItems(self.locations)
        filter_layout.addRow("Location:", location_dropdown)
        from_date_dropdown = QDateEdit(self)
        from_date_dropdown.setCalendarPopup(True)
        filter_layout.addRow("From Date:", from_date_dropdown)
        to_date_dropdown = QDateEdit(self)
        to_date_dropdown.setCalendarPopup(True)
        filter_layout.addRow("To Date:", to_date_dropdown)
        apply_button = QPushButton("Apply", self)
        apply_button.clicked.connect(lambda: self.apply_filters({
            'engineer': engineer_dropdown.currentText(),
            'location': location_dropdown.currentText(),
            'from_date': from_date_dropdown.date().toString('yyyy/MM/dd'),
            'to_date': to_date_dropdown.date().toString('yyyy/MM/dd') if not to_date_dropdown.date().isNull() else None,
        }))
        filter_layout.addRow(apply_button)
        self.layout.addWidget(filter_options_widget)
    
    def display_data(self, data):
            """
            This method clears table and populates it with the provided data.
            It sets up the heades,aligns the text,and applies styling to the table.
            Args:
                data(list): A list of dictionaries containing booking data.
            Raises:
                None
            Returns:
                None
            """
            self.table.clear()
            if not data or not data[0]:
                return
            self.table.setRowCount(len(data))
            self.table.setColumnCount(len(data[0]) + 1)  # Add one extra column for "Remark"
            header_labels = list(data[0].keys()) + ["Remark"] 
            if 'Created Date' not in header_labels:
                header_labels.append('Created Date')
            self.table.setHorizontalHeaderLabels(header_labels)
            for j, key in enumerate(header_labels):
                self.table.setHorizontalHeaderItem(j, QTableWidgetItem(key))
            for i, row in enumerate(data):
                for j, key in enumerate(header_labels):
                    if key == "Remark":  
                        item = QTableWidgetItem(row.get(key, ''))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() | Qt.ItemIsEditable) 
                    else:
                        item = QTableWidgetItem(str(row.get(key, '')))
                        if key == 'engineer' or key == 'location' or key == 'date' or key == 'duration' or key == 'Created Date':
                            item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(i, j, item)
            for j in range(self.table.columnCount()):
                self.table.setSortingEnabled(True)
                # Apply style sheet to set background color for specific columns and header
            self.table.setStyleSheet(
        "QHeaderView::section {"
        "   background-color: yellow;"
        "   color: black;"  # Set text color to black
        "}"
        "QTableWidget::item {"
        "   background-color: #f5f5f5;"
        "}"
        "QTableWidget {"
        "   alternate-background-color: #f5f5f5;"
        "   gridline-color: black;"  # Set grid line color to black
        "}"
    )
    def export_to_csv(self):
            """
            This methods allows the user to select a file name and format for exporting 
            the table data.It then writes the data to selected files in either csv or 
            excel format.
            Args:
                self: Instance of a class.
            Raises:
                None
            Returns:
                None
            """
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(self, "Save Report", "", "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)", options=options)

            if file_name:
                if file_name.endswith('.csv'):
                    with open(file_name, 'w', newline='') as csvfile:
                        csv_writer = csv.writer(csvfile)
                        
                        # Write header
                        header_labels = [self.table.horizontalHeaderItem(j).text() for j in range(self.table.columnCount())]
                        csv_writer.writerow(header_labels)

                        # Write data
                        for i in range(self.table.rowCount()):
                            row_data = [self.table.item(i, j).text() for j in range(self.table.columnCount())]
                            csv_writer.writerow(row_data)
                elif file_name.endswith('.xlsx'):
                    workbook = xlsxwriter.Workbook(file_name)
                    worksheet = workbook.add_worksheet()
                    # Write header
                    header_labels = [self.table.horizontalHeaderItem(j).text() for j in range(self.table.columnCount())]
                    for j, label in enumerate(header_labels):
                        worksheet.write(0, j, label)
                    # Write data
                    for i in range(self.table.rowCount()):
                        for j in range(self.table.columnCount()):
                            worksheet.write(i + 1, j, self.table.item(i, j).text())
                    workbook.close()
    def filter_table(self):
        """
        This method is responsible for filtering data displayed in the table based on
        selected criteria such engineer,location and date range.
        Args:
            self: Instance of class
        Raises:
            None
        Returns:
            None
        """
        pass
    
    def apply_filters(self):
        """
        This method retrieves the selected filters from the filter widgets,constructs
        a dictionary containing filter criteria,and then calls the 'filter_data' to 
        filter the data accordingly.The filtered data is then displayed in the table 
        using 'display_data' method.
        Args:
            self: Instance of a class
        Raises:
            None
        Returns:
            None
        """
        from_date_dropdown = self.from_date_edit.date().toString('yyyy/MM/dd')
        to_date_dropdown = self.to_date_edit.date().toString('yyyy/MM/dd') if not self.to_date_edit.date().isNull() else None
        filters = {
            'engineer': self.engineer_dropdown.currentText(),
            'location': self.location_dropdown.currentText(),
            'from_date': from_date_dropdown,
            'to_date':  to_date_dropdown,
            'duration': None 
        }
        filtered_data = self.filter_data(filters)
        self.display_data(filtered_data)
    
    def filter_data(self, filters):
        """
        This method filters the data based on criteria specified in the
        'filters' dictionary.
        Args:
            filters(dict): A dictionary containing filters criteria.
        Raises:
            None
        Returns: 
            list:A list of dictionaries containing filtered data.

        """
        if not isinstance(filters, dict):
            return []
        filtered_data = self.data.copy()
        if filters.get('engineer') is not None and filters['engineer'] != "All":
            filtered_data = [entry for entry in filtered_data if entry.get('engineer') == filters['engineer']]
        if filters.get('location') and filters['location'] != "All":
            filtered_data = [entry for entry in filtered_data if entry.get('location') == filters['location']]
        if filters.get('from_date') and filters.get('to_date'):
            # Convert filter date strings to datetime objects
            filter_from_date = dt.datetime.strptime(filters['from_date'], '%Y/%m/%d')
            filter_to_date = dt.datetime.strptime(filters['to_date'], '%Y/%m/%d')
            filtered_data = [
                entry for entry in filtered_data
                if 'date' in entry
                and filter_from_date <= dt.datetime.strptime(entry.get('date', ''), '%Y/%m/%d') <= filter_to_date
            ]
        if filters.get('duration') and filters['duration'] != "All":
            filtered_data = [entry for entry in filtered_data if entry.get('duration') == int(filters['duration'].split()[0])]
        return filtered_data

    def get_available_dates(self):
        """
        This method extracts and returns the unique data present in the data.
        Args:
            self: Instance of class.
        Raises:
            None
        Returns:
            list: A list of unique dates.
        """
        return list(set(entry['date'] for entry in self.data))
    
    def add_booked_data(self, booked_data):
        """
        This method appends the provided booked data to the existing data and
        updates the table to display the updated data.
        Args:
            booked_data(dict): A dictionary containing booked data to be added.
        Raises:
            None
        Returns:
            None
        """
        self.data.append(booked_data)
        self.display_data(self.data)

class OpenUrlThread(QThread):
    """
    This class allows opening a URL asynchronously without blocking the main thread.
    """
    finished = pyqtSignal()
    
    def __init__(self, url):
        """
        Initialize the OpenUrlThread with the URL to be opened.
        Args:
            URL(str):The URL to be opened asynchronously.
        Raises:
            None
        Returns:
            None
        """
        super().__init__()
        self.url = url
    
    def run(self):
        """
        This method is called when the thread stars.It opens the URL specified during 
        initialization asynchronously using QDesktopServices.open URl.After opening 
        the URL it emits the 'finished' signal.
        Args:
            self: Instance of the class.
        Raises:
            None
        
        """
        QDesktopServices.openUrl(QUrl(self.url))
        self.finished.emit()

class CustomTableWidget(QTableWidget):
    """
    A custom table widget class.
    """
    
    def __init__(self, parent=None):
        """
        This method initializes the CustomTableWidget.
        Args:
            parent(QWidget): The parent widget to which this table widget belongs.Default is None.
        Raises:
            None
        Returns:
            None
        """
        super().__init__(parent)
        self.parent = parent
        self.cellChanged.connect(self.save_data_on_cell_change)
        self.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.selectionModel().selectionChanged.connect(self.reset_highlighted_cells)

    def reset_highlighted_cells(self, selected, deselected):
        for index in deselected.indexes():
            item = self.itemFromIndex(index)
            if item:
                item.setBackground(QColor(Qt.white))  # Reset background color

    
    def save_data_on_cell_change(self, row, column):
        """
            Retrieves the value of the cell at the specified row and column. If the column corresponds
            to the "Site Address" column, it extracts the location from the cell text. If the location
            is not empty, it saves the location to JSON.
            Args:
                row (int): The row index of the changed cell.
                column (int): The column index of the changed cell.
            Raises:
                None
            Returns:
                None
            """
        if column == self.get_column_index("Site Address"):
                location = self.item(row, column).text()
                if location.strip():  
                    self.save_location_to_json(location)

    def save_location_to_json(self, location):
        """
        Tries to load existing data from "booked_slot.json". If the file is not found, initializes
        data with an empty dictionary. Checks if the given location is already present in the data.
        If not, adds the location to the "locations" dictionary in the data. Finally, writes the
        updated data back to "booked_slot.json".
        Args:
            location (str): The location to be saved to the JSON file.
        Raises:
            FileNotFoundError: If the "booked_slot.json" file is not found.
        Returns:
            None

        """
        try:
            with open("booked_slot.json", "r") as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            data = {"locations": {}}

        if location not in data["locations"]:
            data["locations"][location] = []

        with open("booked_slot.json", "w") as json_file:
            json.dump(data, json_file, indent=4)

    def get_column_index(self, column_name):
        """
        Iterates over the columns in the header to find the index of the column with the specified name.
        If the column is found, its index is returned. If not found, returns -1.
        Args:
            column_name (str): The name of the column to retrieve the index for.
        Raises:
            None
        Returns:
            int: The index of the column, or -1 if not found.


        """
        header = self.horizontalHeader()
        model = header.model()
        for index in range(model.columnCount()):
            if model.headerData(index, Qt.Horizontal) == column_name:
                return index
        return -1   

    def mouseDoubleClickEvent(self, event):
        """
        This method is called when the user double-clicks on table widget.It checks
        if the double click occured on valid cell and makes the cell editable if it
        is.Then it calls the base class method to handle the event further.
        Args:
            event(QMouseEvent): The mouse event that triggered the double-click.
        Raises:
            None
        Returns:
            None
        """
        index = self.indexAt(event.pos())
        if index.isValid():
            self.edit(index)  
        super().mouseDoubleClickEvent(event)
    book_slot_requested = pyqtSignal()
                                    
    def open_url(self):
        """
        Retrieves the currently selected item in the table view. If the item contains a URL, 
        it opens the URL in the default web browser using the `open_google` method.
        Args:
            self
        Raises:
            None
        Returns:
            None

        """
        index = self.currentIndex()
        if index.isValid():
            item = self.item(index.row(), index.column())
            if item is not None:
                url = item.data(Qt.TextDecorationRole)
                if url:
                    self.open_google()

    def delete_row(self):
        """
        Retrieves the indices of the selected rows using the selection model. Then, iterates
        through the selected rows in reverse order and removes them from the table view.
        Args:
            Self
        Raises:
            None
        Returns:
            None
        """
        selected_rows = set(index.row() for index in self.selectionModel().selectedRows())
        for row in sorted(selected_rows, reverse=True):
            self.removeRow(row)    
    
    def mousePressEvent(self, event):
        """
        This method is called when the user presses a mouse button within table widget.
        It checks which mouse button is pressed and performs actions accordingly. If 
        the Right  buttton is pressed,it checks if the click occurred on any cell and
        opens a context menu to book a slot or make cell editable.If the  button is
        pressed,it checks if the click occurred on the "COMPLAINT NO." columns and opens 
        context menu to open URL.
        Args:
            event(QMouseEvent):The mouse events that triggered the press.
        Raises:
            None
        Returns:
            None
        """
        item = self.itemAt(event.pos())
        if item:
            item.setSelected(True)  # Highlight the clicked cell
            item.setBackground(QColor(Qt.blue))  # Set background color to blue
        super().mousePressEvent(event)
        global book_from_complaint # Define the global variable
        if event.button() == Qt.RightButton:
            index = self.indexAt(event.pos())
            if index.isValid():
                menu = QMenu(self)
                insert_row_action = menu.addAction("Insert Row")
                delete_row_action = menu.addAction("Delete Row")
                book_slot_action = menu.addAction("Book Slot")
                open_url_action = menu.addAction("Open URL")
             
                action = menu.exec_(self.viewport().mapToGlobal(event.pos()))
                if action == insert_row_action:
                    self.insertRow(index.row())  
                elif action == delete_row_action:
                    self.delete_row()       
                elif action == book_slot_action:
                    row = index.row()
                    print("Book Slot action clicked in row:", row)
                    self.open_time_slot_dialog()
                    self.complaint_book_slot(row)
                    # book_from_complaint =True
                elif action == open_url_action:
                    self.open_google()  
        else:
            pass

  

    def complaint_book_slot(self, row,current_slot=None):
        """
        Retrieves the relevant information from the specified row in the table view, such as the assigned date,
        site address, and engineer. This information is then used to create a slot object for booking. If a 
        current slot object is provided, it is updated with the retrieved information. 
        Args:
            row (int): The index of the row from which to extract information for booking.
            current_slot (dict, optional): The current slot object. Defaults to None.
        Raises:
            None
        Returns:
            None
        """
        print("Book Slot action clicked in row:", row)
        with open("settings.json", "r") as settings_file:
            settings_data = json.load(settings_file)
        assigned_date_item = self.item(row, self.get_column_index(settings_data["complaint"]["booking_mapping"]["date"]))
        site_address_item = self.item(row, self.get_column_index(settings_data["complaint"]["booking_mapping"]["location"]))
        engineer_item = self.item(row, self.get_column_index(settings_data["complaint"]["booking_mapping"]["engineer"]))

        assigned_date_str = assigned_date_item.text().strip() if assigned_date_item else ""
        site_address = site_address_item.text().strip() if site_address_item else ""
        engineer = engineer_item.text().strip() if engineer_item else ""
        default_duration = settings_data.get("calendar", {}).get("default_duration")
        print("Assigned Date:", assigned_date_str)
        print("Site Address:", site_address)
        print("Engineer:", engineer)
        current_slot = {
            "date": assigned_date_str,
            "time": default_duration,  
            "engineer": engineer,
            "location": site_address
        }
        print(current_slot)
        self.parent.book_slot(current_slot)

    def open_time_slot_dialog(self):
        """
        This method opens the booking dialog if available.
        """ 
        parent_widget = self.parentWidget()
        while parent_widget:
            if isinstance(parent_widget, QTabWidget):
                parent_widget.setCurrentIndex(0)
                calendar_widget = parent_widget.currentWidget()
                if isinstance(calendar_widget, AppointmentSchedulerApp):
                    calendar_widget.show_book_slot_dialog()
                    break
            parent_widget = parent_widget.parentWidget()
        else:
            print("Error: Parent widget of type QTabWidget not found.")
   
    def set_hyperlink_item(self, row, column, text, url):
        """
        This method creates a QTableWidgetItem with specified text and set its text color to
        blue to indicate it's a hyperlink.It sets data of the item to provided URL,so clicking
        on item will open the URL.The item ia aligned to the center and then set in the table
        at specidied row and column.
        Args:
            row(int): The row index where item will be set.
            column(int): The column index where item will be set.
            text(str): The text to display in the item.
            url(str): The URL associated with the item.
        Raises:
            None
        Returns:
            None
        """
        item = QTableWidgetItem(text)
        item.setData(Qt.TextColorRole, Qt.blue)  
        item.setData(Qt.TextDecorationRole, QUrl(url)) 
        item.setTextAlignment(Qt.AlignCenter) 
        self.setItem(row, column, item)
    
    def open_google(self):
        """
        This method uses QDesktopServices to open the specified URL,which opens the 
        webpage of the corresponding company url  directly from the complaint tab .
        Args:
            self: Instance of a class.
        Raises:
            None
        Returns: 
            None
        """
        QDesktopServices.openUrl(QUrl("https://daksh.teamerge.in/admin/ticket_tool/ticket_detail"))

class ComplaintTab(QWidget):
    """
    This class contains functionality for managing complaints including displaying
    a table of complaints,filtering complaints,and performing actions such as 
    booking slots and opening URL.
    """
    book_slot_requested = pyqtSignal()
    def __init__(self, parent:QMainWindow):
        """
        This method initializes the ComplaintTab Widget
        Args:
            self: Instance of class.
        Raises:
            None
        Returns:
            None
        """
        super().__init__(parent)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_excel_file)
        self.parent = parent
        # print("This Panel : ", self.parent)
        self.file_name = None 
        self.checkbox_states = {}
        self.table = CustomTableWidget(parent)
        self.table.setShowGrid(True) 
        self.table.setColumnCount(len(self.header_labels))
        self.table.setHorizontalHeaderLabels(self.header_labels)
        num_blank_rows = 70  
        self.table.setRowCount(num_blank_rows)
        self.table.setStyleSheet(
            "QHeaderView::section {"
            "   background-color: yellow;"
            "}"
        )
        # Main layout
        self.layout = QVBoxLayout(self)
        # Group box for filtering elements
        filter_groupbox = QGroupBox("File Menu")                                            
        filter_groupbox_layout = QVBoxLayout(filter_groupbox)
        # Search Filter Edit
        search_filter_edit_layout = QHBoxLayout()
        self.search_filter_edit = QLineEdit()
        self.search_filter_edit.setPlaceholderText("Search...")
        self.search_filter_edit.textChanged.connect(self.filter_table)
        self.search_filter_edit.setStyleSheet("background-color: #ADD8E6;")
        search_filter_edit_layout.addWidget(self.search_filter_edit,stretch=1)
        # Save Button
        self.save_button = QPushButton("Save Excel")
        self.save_button.clicked.connect(self.auto_save_to_excel)
        self.save_button.setStyleSheet("background-color: #ADD8E6;")
        search_filter_edit_layout.addWidget(self.save_button,stretch=1)
        # Download Button
        self.download_button = QPushButton("Download Excel")
        self.download_button.clicked.connect(self.export_to_excel)
        self.download_button.setStyleSheet("background-color: #ADD8E6;")
        search_filter_edit_layout.addWidget(self.download_button,stretch=1)
        # Upload Button
        self.upload_button = QPushButton("Upload Excel")
        self.upload_button.clicked.connect(self.upload_excel)
        self.upload_button.setStyleSheet("background-color: #ADD8E6;")
        search_filter_edit_layout.addWidget(self.upload_button, stretch=1)
        # Setting Button
        self.setting_button = QPushButton()
        self.setting_button.setIcon(QIcon("setting.png"))  # Set icon
        self.setting_button.setIconSize(QSize(16, 16))  # Set icon size
        self.setting_button.setStyleSheet("background-color: #ADD8E6;")
        self.setting_button.clicked.connect(self.open_settings)
        search_filter_edit_layout.addWidget(self.setting_button)
        filter_groupbox_layout.addLayout(search_filter_edit_layout)
        # Add filter group box to main layout
        self.layout.addWidget(filter_groupbox)
        # # Group box for action buttons
        button_groupbox = QGroupBox("Actions")
        button_groupbox_layout = QVBoxLayout(button_groupbox)
        # # Insert Row Button
        insert_button_layout = QHBoxLayout()
        # # # Insert Column Button
        self.insert_column_button = QPushButton("Insert Column")
        self.insert_column_button.clicked.connect(self.insert_column)
        self.insert_column_button.setStyleSheet("background-color: #ADD8E6;")
        insert_button_layout.addWidget(self.insert_column_button)
        # # Delete Column Button
        self.delete_column_button = QPushButton("Delete Column")
        self.delete_column_button.clicked.connect(self.delete_column)
        self.delete_column_button.setStyleSheet("background-color: #ADD8E6;")
        insert_button_layout.addWidget(self.delete_column_button)
        # Create a refresh button with an image
        self.refresh_button = QPushButton()
        self.refresh_button.setIcon(QIcon("download.jpeg"))  
        self.refresh_button.setIconSize(QSize(25, 25)) 
        self.refresh_button.setToolTip("Refresh Table")  
        self.refresh_button.setStyleSheet("background-color: #ADD8E6;")  
        self.refresh_button.setFixedSize(30, 25)
        insert_button_layout.addWidget(self.refresh_button)
        # Connect the button's clicked signal to the refresh_table method
        self.refresh_button.clicked.connect(self.refresh_table)
        button_groupbox_layout.addLayout(insert_button_layout)
        self.layout.addWidget(button_groupbox)
        self.layout.addWidget(self.table)
        # Styling for Filter Group Box
        filter_groupbox.setStyleSheet("QGroupBox { font-size: 16px; \
                                 border: 2px solid black; \
                                 border-radius: 10px; \
                                 padding-top: 20px; \
                                 background-color: #f0f0f0; \
                                 } \
                                 QGroupBox::title { subcontrol-origin: margin; \
                                subcontrol-position: top center; \
                                 padding: 0 3px; \
                                 color: black; \
                                 font-weight: bold; \
                                 }")
        # # Styling for Button Group Box
        # button_groupbox.setStyleSheet("QGroupBox { font-size: 16px; \
        #                           border: 2px solid black; \
        #                           border-radius: 10px; \
        #                           padding-top: 20px; \
        #                           background-color: #f0f0f0; \
        #                           } \
        #                           QGroupBox::title { subcontrol-origin: margin; \
        #                           subcontrol-position: top center; \
        #                           padding: 0 3px; \
        #                           color: black; \
        #                           font-weight: bold; \
        #    
        #                        }")

    def showEvent(self, event):
        # Start timer when the widget is shown
        self.start_timer()
    def start_timer(self):
        # Set timer interval in milliseconds (e.g., 5000 ms for 5 seconds)
        self.timer.start(1000)  
    def stop_timer(self):
        self.timer.stop()
    def load_excel_file(self):
        file_name = r"C:\Users\HP\OneDrive\Desktop\aa.xlsx"
        print("Loading Excel file:", file_name)
        try:
            workbook = openpyxl.load_workbook(file_name, read_only=True)
            sheet = workbook.active
            self.table.clearContents()
            self.table.setRowCount(0)
            for row_data in sheet.iter_rows(values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                # Iterate over columns in the row
                for col_position, value in enumerate(row_data):
                    if value is None or (isinstance(value, float) and math.isnan(value)):
                        value = ''
                    item = QTableWidgetItem(str(value))  # Convert value to string
                    self.table.setItem(row_position, col_position, item)
            
            print("Excel data loaded successfully!")
        except FileNotFoundError:
            print("Error: File not found.")
        except Exception as e:
            print(f"Error occurred while loading Excel file: {e}")
        self.stop_timer() 
    def save_column_names(self, header_labels):
        """
        This method save column names to the setting file.
        Args:
            header_labels(list of str):The list of column names to be saved.
        Raises:
            FileNotFoundError: If the setting file is not  found.
            Exception: If an error occurs while saving column names.
        Returns:
            None
        """
        try:
            with open('settings.json', 'r+') as file:
                settings_data = json.load(file)
                settings_data["complaint"]["columns"] = header_labels
                file.seek(0)  
                json.dump(settings_data, file, indent=4) 
                file.truncate()  
        except FileNotFoundError:
            print("Settings file not found.")
        except Exception as e:
            print(f"Error occurred while saving column names: {e}")
    
    def set_hyperlink_item(self, row, column, text, url):
        """
        This method sets a hyperlink item in the specified rows and columns.
        Args:
            row(int): The row index.
            column(int): The column index.
            text(str):The text to display.
            url(str): URL to be opened when clicked.
        Raises:
            None
        Returns:
            None
        """
        item = QTableWidgetItem(text)
        item.setData(Qt.TextColorRole, Qt.blue)  # Set text color to blue
        item.setData(Qt.DecorationRole, QUrl(url))  # Set decoration role as QUrl
        item.setTextAlignment(Qt.AlignCenter)  # Align text to center
        self.setItem(row, column, item)
    
    def reset_complaint_tab(self):
        """
        This method reset the complaint tab by showing all columns and clearing
        any applied text filter.
        Args:
            self: Instance of a class
        Raises:
            None
        Returns:
            None
        """
        # Show all columns in the complaint tab
        for column in range(self.table.columnCount()):
            self.table.setColumnHidden(column, False)
        # Clear any applied text filter
        self.search_filter_edit.clear()
        self.filter_table()  # Apply the filter to show all rows    
    
    def open_settings(self):
        """
        This method reads the settings from 'settings.json' to get column related to 
        complaints.It creates menu with checkboxes for each column and allows users
        to select/deselect columns.The selected column can be applied using 'OK' 
        button and filters can be reset using 'Reset' button.
        Args:
            self: Instance of class.
        Raises:
            None
        Returns:
            None
        """
        with open('settings.json', 'r') as file:
            settings_data = json.load(file)
            complaint_columns = settings_data.get("complaint", {}).get("columns", [])
        # Create a menu to display columns with checkboxes
        menu = QMenu(self.setting_button)
        menu.setStyleSheet("""
            QMenu {
                padding: 10px;
            }
            QMenu::item {
                padding: 5px;
            }
            QPushButton {
                padding: 5px;
            }
        """)
        for column in complaint_columns:
            checkbox_widget = QWidget()
            layout = QHBoxLayout(checkbox_widget)
            checkbox = QCheckBox()
            checkbox.setText(column)
           # Check the checkbox if it was previously selected
            if self.checkbox_states.get(column, False):
                checkbox.setChecked(True)
            layout.addWidget(checkbox)
        # Create a custom action with the checkbox widget
            action = QWidgetAction(menu)
            action.setDefaultWidget(checkbox_widget)
            menu.addAction(action)
        # Add an "OK" button to apply the selected filters
        ok_action = menu.addAction("OK")
        ok_action.triggered.connect(lambda: self.apply_filter(menu))
        reset_action = menu.addAction("Reset")
        reset_action.triggered.connect(lambda: self.reset_filter(menu))
       # Show the menu below the setting button
        menu.exec_(self.setting_button.mapToGlobal(self.setting_button.rect().bottomLeft()))
    
    def reset_filter(self, menu):
        """
        This method reset all checkboxes in the setting menu to their default state.
        Args:
            menu: The QMenu object representing the setting menu.
        Raises:
            None
        Returns:
            None.

        """
        # Reset all checkboxes to default state (unchecked)
        for action in menu.actions():
            if isinstance(action, QWidgetAction):
                widget = action.defaultWidget()
                if widget and isinstance(widget, QWidget):
                    checkbox = widget.findChild(QCheckBox)
                    checkbox.setChecked(False)
        self.checkbox_states = {}
        self.reset_complaint_tab()


    def refresh_table(self):
        """
        This method iterates over each cell in the table and clears the text content of the cell,
        effectively resetting the table to an empty state.
        Args:
            Self
        Raises:
            None
        Returns:
            None

        """
        # Clear the contents of each cell
        for row in range(self.table.rowCount()):
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                if item:
                    item.setText("")  # Clear the text    
    
    def apply_filter(self, menu):
        """
        This method retrieves the selected column names from the checkboxes in 
        the setting menu and store their states.It then shows all rows by default
        and hides column not selected in the checkboxes.Additionally it applies 
        text filter if text is entered in the search filter edit.
        Args:
            menu(Qmenu): The menu containing the filter options.
        Raise:
            None
        Returns:
            None
        """
        # Get the selected column names from the checkboxes
        selected_columns = []
        for action in menu.actions():
            if isinstance(action, QWidgetAction):
                widget = action.defaultWidget()
                if widget and isinstance(widget, QWidget):
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox.isChecked():
                        selected_columns.append(checkbox.text())
                        # Store the state of the checkbox
                        self.checkbox_states[checkbox.text()] = True
                    else:
                        # Unchecked checkboxes should not be stored
                        self.checkbox_states.pop(checkbox.text(), None)
        # Show all rows by default
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
        # Hide columns not selected in the checkboxes
        for column in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(column)
            if header_item.text() not in selected_columns:
                self.table.setColumnHidden(column, True)
            else:
                self.table.setColumnHidden(column, False)
        # Apply text filter if needed
        text = self.search_filter_edit.text()
        if text:
            self.filter_table(text)
    
    def open_url(self):
        """
        This method opens the URL associated with the currently selected items in 
        view,if available.
        Args:
            self: Instance of class.
        Raises:
            None
        Returns:
            None
        """
        index = self.currentIndex()
        if index.isValid():
            item = self.item(index.row(), index.column())
            if item is not None:
                url = item.data(Qt.DecorationRole).toString()
                if url:
                    QDesktopServices.openUrl(QUrl(url))
    
    def upload_excel(self):
        """
        This method opens a file dialog to allow the users to select an Excel file.
        Clears existing data in the table,loads the data from selected excel file 
        and sets the file name attribute.
        Args:
            self: Instance of a class.
        Raises:
            None
        Returns:
            None
        """
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Upload Excel File", "", "Excel Files (*.xlsx *.xls);;All Files (*)", options=options)
        if file_name:
            self.table.clear()  # Clear existing data in the table
            self.load_data(file_name)
            self.file_name = file_name  

    def header_labels(settings_file):
        """
        This method reads the specified setting file containing JSON data.Retrieves   
        the list of column labels for complaint data. If the setting file is not 
        found an empty list is returned.
        Args:
            setting_file(str): The path to the setting file.
        Raises:
            FileNotFoundError: If the setting file is not found.
        Returns:
            list: A list of column labels for complaint data or an empty list,if the
            setting file not found.
        """
        try:
            with open(settings_file, "r") as f:
                settings_data = json.load(f)
                complaint_columns = settings_data.get("complaint", {}).get("columns", [])
                return complaint_columns
        except FileNotFoundError:
            print("Settings file not found.")
            return []
    # Usage
    def load_header_labels(self):
        """
        This method reads the header labels from the specified settings file and prints them to the console.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        settings_file = "settings.json"
        header_labels = self.header_labels(settings_file)
        print(header_labels)
    # Usage
    settings_file = "settings.json"
    header_labels = header_labels(settings_file)
    print(header_labels)
    
    def auto_save_to_excel(self):
        """
        This method attempts to auto-save the data from the table to the Excel file
        specified by 'self.file_name'.If successful,prints a success message.
        Args:
            self: Instance of class
        Raises:
            FileNotFoundError: If the excel file specified by 'self.file_name' not found.
            PermissionError: If there is a permission issue while accessing the excel file.
            Exception: For other unexpected error.
        Returns:
            None
        """
        try:
            if self.file_name:
                file_name = self.file_name
                workbook = load_workbook(file_name)
                sheet = workbook.active
                # Clear existing data in the sheet
                sheet.delete_rows(2, sheet.max_row)
                # Update the entire table data to the Excel sheet
                for i in range(self.table.rowCount()):
                    for j in range(self.table.columnCount()):
                        item = self.table.item(i, j)
                        value = item.text() if item is not None else ''
                        sheet.cell(row=i + 2, column=j + 1, value=value)
                # Save the changes back to the file
                workbook.save(file_name)
                print(f"Auto-save successful to {file_name}")
        except FileNotFoundError as e:
            print(f"Error: File not found: {e}")
        except PermissionError as e:
            print(f"Error: Permission issue: {e}")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print(f"Error: Invalid Excel file: {e}")
        except Exception as e:
            print(f"Unexpected error in auto_save_to_excel: {e}")
                                                  
    def filter_table(self, text=None):
        """
        This method filters the table to show only rows where any cell contains the 
        specified text.If no text is provided,it retrieves the text from search filter
        edit widget.
        Args:
            text(str,optional): The text to filter the table with. If not provided,the 
            text is retrieved from the search filter edit widget.
        Raises:
            None
        Returns:
            None
        """
        if text is None:
            text = self.search_filter_edit.text()            
        for row in range(self.table.rowCount()):
            row_hidden = True
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                if item is not None and text.lower() in item.text().lower():
                    row_hidden = False
                    break
            self.table.setRowHidden(row, row_hidden)
            
    def insert_row(self):
        """
        This method insert a new row at the end of the table.After inertion it automatically saves
        the changes to the associated excel file.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        current_rows = self.table.rowCount()
        self.table.insertRow(current_rows)
        self.auto_save_to_excel()

    def insert_column(self):
        """
        This method inserts new column at the bottom of the table.After insertion it automatically
        saves then changes to the associated excel file.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        current_columns = self.table.columnCount()
        header_label, ok = QInputDialog.getText(self, "Input Dialog", "Enter column header:")
        if ok:
            self.table.setColumnCount(current_columns + 1)
            self.table.setHorizontalHeaderItem(current_columns, QTableWidgetItem(header_label))
            # Update header_labels list
            self.header_labels.append(header_label)
            # Save updated column names to file
            self.save_column_names(self.header_labels)
            # Set the horizontal header labels again
            self.table.setHorizontalHeaderLabels(self.header_labels)
        self.auto_save_to_excel()

    def save_column_names(self, header_labels):
        """
        This method updates the column name for complaints in the setting file with 
        the provided ;header_labels' list.
        Args:
            header_labels(list): A list of column names to be saved to the setting file.
        Raises:
            FileNotFoundError: If the settings file is not found.
            Exception: For other unexpected error.
        Returns:
            None

        """
        settings_file = "settings.json"
        try:
            with open(settings_file, "r") as f:
                settings_data = json.load(f)
                settings_data["complaint"]["columns"] = header_labels
            with open(settings_file, "w") as f:
                json.dump(settings_data, f, indent=4)
            print("Column names updated successfully.")
        except FileNotFoundError:
            print("Settings file not found.")
        except Exception as e:
            print(f"Error while saving column names: {e}")

    def delete_row(self):
        """
        Deletes the rows that are currently selected in the table.After deletion it automatically saves 
        the changes to the excel fie.
        Args:
            self
        Raises:
            None
        Returns:
            None

        """
        selected_rows = set(index.row() for index in self.table.selectionModel().selectedRows())
        for row in sorted(selected_rows, reverse=True):
            self.table.removeRow(row)
        self.auto_save_to_excel()

    def delete_column(self):
        """
        This method deletes the column that are currently selected in the table.After deletion it automatically 
        saves the changes to excel files.
        Args:
            self
        Raises:
            Exception: If there is an error removing column data from the file.
        Returns:
            None
        """
        try:
          
            selected_columns = [index.column() for index in self.table.selectionModel().selectedColumns()]
            if self.file_name:
               
                workbook = openpyxl.load_workbook(self.file_name)
                sheet = workbook.active
               
                for col in sorted(selected_columns, reverse=True):
                    sheet.delete_cols(col + 1)  
               
                workbook.save(self.file_name)
                print(f"Column data removed successfully from {self.file_name}")
           
            new_header_labels = [label for col, label in enumerate(self.header_labels) if col not in selected_columns]
            self.header_labels = new_header_labels
            self.table.setColumnCount(len(self.header_labels))  
            self.table.setHorizontalHeaderLabels(self.header_labels)
          
            self.save_column_names(self.header_labels)
           
            self.auto_save_to_excel()
        except Exception as e:
            print(f"Error removing column data from file: {e}")
        
    def load_data(self, file_name, existing_header_labels=None, new_header_labels=None):
        """
        This method reads data from the specified Excel file and loads it into a table widget. 
        It allows for mapping column headers from an existing format to a new format.
        Args:
            file_name (str): The name of the Excel file to load.
            existing_header_labels (list of str, optional): The existing column headers in the table widget. Defaults to None.
            new_header_labels (list of str, optional): The new column headers to map the data to. Defaults to None.
        Raises:
            FileNotFoundError: If the specified Excel file cannot be found.
        Returns:
            None


        """
        try:
            workbook = openpyxl.load_workbook(file_name, read_only=True)
            sheet = workbook.active
            self.table.clearContents()
            self.table.setRowCount(0)
            loaded_data = []
            is_first_row = True
            for row_data in sheet.iter_rows(values_only=True):
                if is_first_row:
                    is_first_row = False
                    if existing_header_labels and new_header_labels and existing_header_labels != new_header_labels:
                        mapping = {"OldColumnName": "NewColumnName"}
                        continue
                    else:
                        header_labels = [str(value) for value in row_data]
                        continue
                if existing_header_labels and new_header_labels and existing_header_labels != new_header_labels:
                    mapped_row_data = [None] * len(existing_header_labels)
                    for col_position, value in enumerate(row_data):
                        column_name = new_header_labels[col_position]
                        if column_name in existing_header_labels:
                            mapped_column_index = existing_header_labels.index(mapping.get(column_name, column_name))
                            mapped_row_data[mapped_column_index] = value
                    loaded_data.append(mapped_row_data)
                else:
                    loaded_data.append(row_data)
            
            for row_data in loaded_data:
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for col_position, value in enumerate(row_data):
                    if value is None or (isinstance(value, float) and math.isnan(value)):
                        value = ''
                    item = QTableWidgetItem()
                   
                    if header_labels[col_position] == "COMPLAINT NO.":
                        font = QFont()
                        font.setUnderline(True)
                        font.setPointSize(10)
                        item.setFont(font)
                        item.setForeground(QColor(Qt.blue))
                        item.setData(Qt.DisplayRole, value)
                        item.setData(Qt.UserRole, value)
                    else:
                        item.setData(Qt.DisplayRole, value)
                    self.table.setItem(row_position, col_position, item)
            
            if loaded_data:
                self.table.setColumnCount(len(loaded_data[0]))
            
            if not existing_header_labels or existing_header_labels == new_header_labels:
                self.table.setHorizontalHeaderLabels(header_labels)
            else:
                self.table.setHorizontalHeaderLabels(existing_header_labels)
            
            print("Excel data loaded successfully!")
        except FileNotFoundError:
            print(f"Error: File not found: {file_name}")
        except Exception as e:
            print(f"Error reading Excel file: {e}")
        
    def export_to_excel(self):
        """
        This method allows the users to save data from table widgets to an Excel file.
        Arg:
            self
        Raises:
            None
        Returns:
            None
        """
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Complaints to Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            header_labels = [self.table.horizontalHeaderItem(j).text() for j in range(self.table.columnCount())]
            sheet.append(header_labels)
            for i in range(self.table.rowCount()):
                row_data = [self.table.item(i, j).text() if self.table.item(i, j) is not None else '' for j in range(self.table.columnCount())]
                sheet.append(row_data)
            workbook.save(file_name)

class AppointmentSchedulerApp(QMainWindow):
    """
    Main application window for the Appointment Scheduler.
    """
    def __init__(self,parent=None):
       
        """
        This method sets up the main window with various widgets,actions and tabs for Appointment
        Scheduler Application.Load data from JSON files,initialzes times for automatic data refreshing,
        and handles users authentication.
        Args:
            None
        Raises:
            None 
        Returns:
            None
        """
        super().__init__()
        self.parent= parent
        self.setWindowTitle("Appointment Scheduler")
        self.setGeometry(100, 100, 1200, 800)
        self.central_widget = QTabWidget(self)
        self.setCentralWidget(self.central_widget)
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)
        refresh_button = QPushButton("Refresh Data", self)
        refresh_button.clicked.connect(self.refresh_fetchdata)
        layout.addWidget(refresh_button)
        # File menu
        self.file_menu = self.menuBar().addMenu("File")
        # Add Location action
        self.add_location_action = QAction("Add Location", self)
        self.add_location_action.triggered.connect(self.add_location)
        self.file_menu.addAction(self.add_location_action)
        # Add Engineer action
        self.add_engineer_action = QAction("Add Engineer", self)
        self.add_engineer_action.triggered.connect(self.add_engineer)
        self.file_menu.addAction(self.add_engineer_action)
        # Delete Location action
        self.delete_location_action = QAction("Delete Location", self)
        self.delete_location_action.triggered.connect(self.delete_location)
        self.file_menu.addAction(self.delete_location_action)
        # Delete Engineer action
        self.delete_engineer_action = QAction("Delete Engineer", self)
        self.delete_engineer_action.triggered.connect(self.delete_engineer)
        self.file_menu.addAction(self.delete_engineer_action)
        # Refresh action
        self.refresh_action = QAction(QIcon("refresh button.png"), "Refresh", self)
        self.refresh_action.triggered.connect(self.refresh_fetchdata)
        self.file_menu.addAction(self.refresh_action)
        self.logout_action = QAction(QIcon("logout button.png"), "Logout", self)
        self.logout_action.triggered.connect(self.show_logout_dialog)
        self.file_menu.addAction(self.logout_action) 
        # Calendar tab
        self.calendar_tab = QWidget(self)
        self.central_widget.addTab(self.calendar_tab, "Calendar")
        # Complaint tab
        self.complaint_tab = ComplaintTab(self)
        # self.complaint_tab.book_slot_requested.connect(self.show_book_slot_dialog)
        self.central_widget.addTab(self.complaint_tab, "Complaint")
        self.central_widget.currentChanged.connect(self.tab_changed)
        # Create a vertical layout for the calendar tab
        self.calendar_tab_layout = QVBoxLayout(self.calendar_tab)
        self.calendar_and_slot_layout = QHBoxLayout()
        # Calendar widget
        self.calendar = QCalendarWidget(self)
        self.calendar.clicked.connect(self.show_slot_info)
        self.calendar.setContextMenuPolicy(Qt.CustomContextMenu)
        self.calendar.customContextMenuRequested.connect(self.show_context_menu) 
        self.calendar_and_slot_layout.addWidget(self.calendar)
        # Book slot button
        self.book_button = QPushButton("Book Slot", self)
        self.book_button.clicked.connect(self.show_book_slot_dialog)
        self.book_button.setShortcut("Return")
        # Frame to display assigned engineers
        self.engineers_frame = QFrame(self)
        self.engineers_frame.setFrameShape(QFrame.StyledPanel)
        self.engineers_frame.setFrameShadow(QFrame.Raised)
        self.engineers_frame.setFixedWidth(450)
        # GroupBox inside the Frame
        self.engineers_group_box = QGroupBox("Booked Slots")
        self.engineers_group_box.setStyleSheet("font-weight: bold; font-size: 11px;")
        self.engineers_group_box.setFixedWidth(500)
        self.inner_group_box = QGroupBox("Inner GroupBox")
        self.inner_group_box.setFixedWidth(100)
        self.inner_group_box_layout = QVBoxLayout(self.inner_group_box)
        # Assigned Engineers
        self.assigned_label = QLabel("Assigned Engineers:", self.inner_group_box)
        self.assigned_text_edit = QTextEdit(self.inner_group_box)
        # Unassigned Engineers
        self.unassigned_label = QLabel("Unassigned Engineers:", self.inner_group_box)
        self.unassigned_text_edit = QTextEdit(self.inner_group_box)
        # Add the GroupBox to the Frame
        self.engineers_frame_layout = QVBoxLayout(self.engineers_frame)
        self.engineers_frame_layout.addWidget(self.engineers_group_box)
        self.inner_group_box_layout.addWidget(self.assigned_label)
        self.inner_group_box_layout.addWidget(self.assigned_text_edit)
        self.inner_group_box_layout.addWidget(self.unassigned_label)
        self.inner_group_box_layout.addWidget(self.unassigned_text_edit)
        self.engineers_group_box.setLayout(self.inner_group_box_layout)
        self.calendar_tab_layout.addLayout(self.calendar_and_slot_layout)
        self.calendar_and_slot_layout.addWidget(self.engineers_frame)
        self.calendar_tab_layout.addWidget(self.book_button)
        self.timer = QTimer(self)
        # Timer for resource usage monitoring
        self.resource_timer = QTimer(self)
        self.resource_timer.timeout.connect(self.check_resource_usage)
        self.resource_timer.start(5000)  # Check every 5 seconds
        self.timer.timeout.connect(self.refresh_fetchdata)
        refresh_interval = 10000
        self.timer.start(refresh_interval)
        self.timer.timeout.connect(self.complaint_timer)
        self.start_timer()
        # Data
        self.data = {"bookings":[], "engineers":[], "locations":[]}
        self.data = self.load_json()
        if len(self.data["locations"]) != 0:
            self.locations = {i:[] for i in self.data["locations"]}
        else:
            self.locations = {"Office":[],"Leave":[],"Candor Tikri":[], "DLF Ultima":[], "DLF Primus":[], "Candor Dhundahera":[], "Regal Garden":[]}
            self.save_json("locations", self.locations)
        
        if len(self.data["engineers"]) != 0:
            self.engineers = {i:[] for i in self.data["engineers"]}
        else:
            self.engineers =  {"Deepak":[], "Raman":[], "Shishupal":[], "Vikas":[], "Sachin":[]}
            self.save_json("engineers", self.engineers)
        self.booked_slots = {}
        self.bookings = self.data["bookings"]
        self.dashboard_tab = DashboardWidget(self.locations.keys(), self.engineers.keys(), self)
        self.central_widget.addTab(self.dashboard_tab, "Dashboard")
        self.reports_widget = ReportsWidget(list(self.engineers), list(self.locations), self)
        self.central_widget.addTab(self.reports_widget, "Reports")
        self.central_widget.tabBarClicked.connect(self.handle_tab_click)
        self.central_widget.setCurrentWidget(self.calendar_tab)
        if self.load_existing_credentials():
            self.show()
        else:
            self.login_window = LoginWindow(self)
            self.login_window.login_signal.connect(self.show)
            self.login_window.show()
   
    def tab_changed(self, index):
        if index == 1:  # Check if the Complaint tab is activated
            self.complaint_tab.load_excel_file()

    def check_resource_usage(self):
        """
        Check Appointment Scheduler application's CPU and memory usage and show alert if it exceeds thresholds.

        Returns:
            None
        """
        try:
           
            pid = psutil.Process().pid
            print("Current Process ID:", pid)
            app_cpu_percent = psutil.Process(pid).cpu_percent()
            print("Current CPU Usage:", app_cpu_percent, "%")
            app_memory_usage = psutil.Process(pid).memory_info().rss / (1024 * 1024)  # Convert to MB
            print("Current Memory Usage:", app_memory_usage, "MB")
            if app_cpu_percent >= 5 and app_memory_usage >= 200:
                print("High resource usage detected. Showing alert...")
                self.show_resource_alert(app_cpu_percent, app_memory_usage)
            else:
                print("Resource usage within limits.")
        except Exception as e:
            print("Error checking resource usage:", e)

    def show_resource_alert(self, cpu_percent, memory_usage):
        """
        Show a message box alerting high resource usage.

        Args:
            cpu_percent: CPU usage percentage.
            memory_usage: Memory usage in MB.

        Returns:
            None
        """
        try:
            msg_box = QMessageBox()
            msg_box.setWindowTitle('Resource Alert')
            msg_box.setText(f'System will shut down the Appointment Scheduler due to high resource usage.\n\n'
                            f'CPU Usage: {cpu_percent}%\n'
                            f'Memory Usage: {memory_usage} MB')
            msg_box.exec_()
            self.close()  # Shut down the application
        except Exception as e:
            print("Error displaying resource alert:", e)
    def handle_tab_click(self, index):
        """
        This methdod handles the click events on the tabs of the main application window.If 
        the click tab is labeled "Logout",it invokes the method to show the logout dialog.
        Args:
            index(int): The index of the clicked tab.
        Raises:
            None
        Returns:
            None
        """
        current_tab_text = self.central_widget.tabText(index)
        if current_tab_text == "Logout":
            self.show_logout_dialog()

    def show_logout_dialog(self):
        """
        This method displays a dialog box to confirm whether user wants to logout.
        The dialog box contains options to confirm or cancel the logout action.
        Args:
            Self
        Raises:
            None
        Returns:
            None
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("Logout")
        dialog.setGeometry(500, 300, 300, 100)
        layout = QVBoxLayout(dialog)
        confirmation_label = QLabel("Do you want to logout?", dialog)
        layout.addWidget(confirmation_label)
        button_layout = QHBoxLayout()
        ok_button = QPushButton("OK", dialog)
        ok_button.clicked.connect(self.logout)
        button_layout.addWidget(ok_button)
        cancel_button = QPushButton("Cancel", dialog)
        cancel_button.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        dialog.exec_()

    def logout(self):
        """
        This method updates the status of the current user to "inactive" in the liscence
        file.Closes the application window after the logout process.
        Args:
            self
        Raise:
            None
        Returns:
            None
        """
        try:
            with open("license.json", "r") as license_file:
                existing_data = json.load(license_file)
                for user in existing_data.get("user", []):
                    user["status"] = "inactive"
            with open("license.json", "w") as license_file:
                json.dump(existing_data, license_file, indent=4)
            self.close()
        except FileNotFoundError:
            QMessageBox.warning(self, "Logout Error", "License file not found. Logout failed.")
    
    def load_existing_credentials(self):
        """
        This method checks if there are active user credentials in the license file.If active credentials
        are found it returns "True",else it returns "False".
        Args:
            self
        Raises:
            FileNotFoundError:If the license file is not found.

        Returns:
            bool:True if active user credentials are found,False otherwise.
        """
        try:
            with open("license.json", "r") as license_file:
                existing_data = json.load(license_file)
                for i in existing_data["user"]:
                    if i["status"] == "active":
                        return True
                return False
        except FileNotFoundError:
            return False

    def update_calendar_colors(self):
        """
        Calculates the total number of assigned engineers for each date and updates the color
        of the corresponding calendar date cells accordingly.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        # Calculate the total number of assigned engineers for each date
        engineer_counts = {}
        max_count = 0
        for booking in self.bookings:
            date = QDate.fromString(booking["date"], "yyyy/MM/dd")
            if date not in engineer_counts:
                engineer_counts[date] = 0
            engineer_counts[date] += 1
            max_count = max(max_count, engineer_counts[date])

        # Define start and end colors for the gradient
        start_color = QColor(204, 255, 204)  # Light green
        end_color = QColor(0, 102, 0)        # Dark green

        for date, count in engineer_counts.items():
            # Calculate the gradient color based on the count of assigned engineers
            gradient = count / max_count
            red = start_color.red() + int(gradient * (end_color.red() - start_color.red()))
            green = start_color.green() + int(gradient * (end_color.green() - start_color.green()))
            blue = start_color.blue() + int(gradient * (end_color.blue() - start_color.blue()))

            cell_color = QColor(red, green, blue)

            # Set the background color format for the date cell
            date_format = QTextCharFormat()
            date_format.setBackground(cell_color)

            # Set the text color to contrast with the background
            text_color = QColor(255 - cell_color.red(), 255 - cell_color.green(), 255 - cell_color.blue())
            date_format.setForeground(text_color)

            # Set the format for the entire cell
            self.calendar.setDateTextFormat(date, date_format)
        # Refresh the calendar display
        self.calendar.repaint()

    def show_context_menu(self, pos):
        """
        This method displays a context menu when triggered by a right-click event on the calendar.
        If bookings exist for selected date,it allows the user to edit the booking through a dialog
        window.
        Args:
            pos(QPoint):The position if the right click event.
        Raises:
            None
        Returns:
            None
        """
        selected_date = self.calendar.selectedDate().toString('yyyy/MM/dd')
        filtered_bookings = [booking for booking in self.bookings if booking["date"] == selected_date]
        if not filtered_bookings:
        # If no bookings, show an error message and return
            error_box = QMessageBox()
            error_box.setIcon(QMessageBox.Warning)
            error_box.setWindowTitle('No Bookings')
            error_box.setText('No bookings found for the selected date.')
            error_box.setStandardButtons(QMessageBox.Ok)
            error_box.exec_()
            return
        context_menu = QMenu(self)
        show_dialog_action = context_menu.addAction("Edit")
        # Connect the "Edit" action to the show_dialog_function with the filtered bookings and selected date as arguments
        show_dialog_action.triggered.connect(lambda: self.show_dialog_function(filtered_bookings, selected_date))
        context_menu.exec_(self.calendar.mapToGlobal(pos))

    def show_dialog_function(self, filtered_bookings, selected_date):
        """
        This method displays a dialog window for editing bookings on the selected date.The 
        dialog window contains a table populated with existing bookings for the date,allowing
        the user to make changes.Upon accepting the changes,the updated bookings are saved.
        Args:
            filtered_bookings(list):A list of booking filtered for the selected date.
            selected_date(str):The selected date in 'yyyy/MM/dd' format.
        Raises:
            None
        Returns:
            None
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Slot")
        dialog.setGeometry(300, 300, 675, 290)
        layout = QVBoxLayout(dialog)
        # Create or update the table directly
        self.populate_table(dialog, filtered_bookings, selected_date)
        # OK button
        ok_button = QPushButton("OK", dialog)
        ok_button.clicked.connect(dialog.accept)
        layout.addWidget(ok_button)
        # Cancel button
        cancel_button = QPushButton("Cancel", dialog)
        cancel_button.clicked.connect(dialog.reject)
        layout.addWidget(cancel_button)
        result = dialog.exec_()
        if result == QDialog.Accepted:
            print("OK button clicked")
            self.save_json("bookings", self.bookings)
        else:
            print("Cancel button clicked")

    def populate_table(self, dialog, filtered_bookings, selected_date):
        """
        This method populates a table in the dialog window with booking information retrieved 
        from the filtered booking list for the selected date.The table includes details such
        as engineer name,location,date,duration,and options to edit or delete bookings.
        Args:
            dialog(QDialog): The dialog window where the table is displayed.
            filtered_booking(list): A list of booking filtered for the selected date.
            selected_date(str): The selected date in 'yyyy/MM/dd' format.
        Raises:
            None
        Returns:
            None
        """
        booking_table_widget = dialog.findChild(QTableWidget, "booking_table_widget")
        if not booking_table_widget:
            booking_table_widget = QTableWidget(dialog)
            booking_table_widget.setObjectName("booking_table_widget")
            layout = dialog.layout()
            layout.addWidget(booking_table_widget)
        booking_table_widget.clearContents()
        booking_table_widget.setRowCount(len(filtered_bookings))
        booking_table_widget.setColumnCount(6)
        booking_table_widget.setHorizontalHeaderLabels(["Engineer", "Location", "Date", "Duration", "Edit", "Delete"])
        
        for row, booking in enumerate(filtered_bookings):
            booking_table_widget.setItem(row, 0, QTableWidgetItem(booking["engineer"]))
            booking_table_widget.setItem(row, 1, QTableWidgetItem(booking["location"]))
            booking_table_widget.setItem(row, 2, QTableWidgetItem(booking["date"]))
            duration_str = f"{booking['duration']} {'Hour' if booking['duration'] == 1 else 'Hours'}"
            booking_table_widget.setItem(row, 3, QTableWidgetItem(duration_str))
            edit_button = QPushButton("Edit", booking_table_widget)
            edit_button.clicked.connect(lambda _, b=booking: self.update_booking(b, dialog))
            booking_table_widget.setCellWidget(row, 4, edit_button)

            delete_button = QPushButton("Delete", booking_table_widget)
            delete_button.clicked.connect(lambda _, b=booking: self.delete_booking(b, dialog))
            booking_table_widget.setCellWidget(row, 5, delete_button)

    def delete_booking(self, booking, parent_dialog):
            """
            This method displays a confirmation dialog to confirm the deletion of a booking for specific 
            engineer on a certain date.If the user confurms the deletion ,a DELETE request is sent to the
            server to delete the booking.Upon successfull deletion,the booking is removed from the local
            booking list,the UI is updated to reflect the changes,and parent dialog window is closed.
            Args:
                booking(dict): The booking to be deleted containing details such as engineer name,
                               location,date and duration.
                parent_dialog(QDialog): The parent dialog window associated with booking.
            Raises:
                request.exceptions.RequestException: If an error occurs while making the DELETE reqyest
                                                     to server.
            Returns:
                None
            """
            confirmation = QMessageBox.question(
                self,
                "Delete Booking",
                f"Do you want to delete the booking for {booking['engineer']} on {booking['date']}?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if confirmation == QMessageBox.Yes:
                delete_api_url = "http://192.168.17.72:5000/delete_booking"
                headers = {"Content-Type": "application/json"}

                delete_data = {
                    "engineer": booking["engineer"],
                    "location": booking["location"],
                    "date": booking["date"],
                    "duration": booking["duration"]
                }
                try:
                    response = requests.delete(delete_api_url, data=json.dumps(delete_data), headers=headers)
                    response.raise_for_status()
                    delete_response = response.json()

                    if "message" in delete_response:
                        print(delete_response["message"])
                        self.bookings.remove(booking)
                        self.save_json("bookings", self.bookings)
                        self.show_slot_info()
                        parent_dialog.accept()
                    else:
                        print("Failed to delete booking. API response:", delete_response)
                        QMessageBox.warning(self, "Delete Failed", "Failed to delete booking. Please try again.")
                except requests.exceptions.RequestException as e:
                    print(f"Error making delete_booking API request: {e}")
                    QMessageBox.critical(self, "Delete Error", f"Error making delete_booking API request: {e}")

    def update_booking(self, booking, parent_dialog):
        """
        This method displays a dialog box window for updating the details of booking,including the engineer,
        location,date and duration.Upon confirming the changes,the booking details are updated locally and
        reflected in the UI.
        Args:
            booking(dict): The booking to be updated containing details such as engineer,name,location and 
                           duration.
            parent_dialog(QDialog): The parent dialog window associated with booking.
        Raises:
            None
        Returns:
            None
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("Update Booking")
        dialog.setGeometry(300, 300, 400, 200)
        layout = QVBoxLayout(dialog)
        engineer_dropdown = QComboBox(dialog)
        engineer_dropdown.addItems(self.engineers.keys())
        engineer_dropdown.setCurrentText(booking["engineer"])
        location_dropdown = QComboBox(dialog)
        location_dropdown.addItems(self.locations.keys())
        location_dropdown.setCurrentText(booking["location"])
        date_edit = QDateEdit(dialog)
        date_edit.setDate(QDate.fromString(booking["date"], "yyyy/MM/dd"))
        duration_spinbox = QSpinBox(dialog)
        duration_spinbox.setMinimum(1)
        duration_spinbox.setMaximum(8)
        duration_spinbox.setValue(booking["duration"])
        ok_button = QPushButton("OK", dialog)
        ok_button.clicked.connect(lambda: self.update_booking_callback(booking, engineer_dropdown, location_dropdown, date_edit, duration_spinbox, dialog, parent_dialog))
        cancel_button = QPushButton("Cancel", dialog)
        cancel_button.clicked.connect(dialog.reject)
        layout.addWidget(QLabel("Engineer:"))
        layout.addWidget(engineer_dropdown)
        layout.addWidget(QLabel("Location:"))
        layout.addWidget(location_dropdown)
        layout.addWidget(QLabel("Date:"))
        layout.addWidget(date_edit)
        layout.addWidget(QLabel("Duration (hours):"))
        layout.addWidget(duration_spinbox)
        layout.addWidget(ok_button)
        layout.addWidget(cancel_button)
        dialog.exec_()

    def update_booking_callback(self, booking, engineer_dropdown, location_dropdown, date_edit, duration_spinbox, dialog, parent_dialog):
        """
        This method updates the details of booking based on user's input from the dialog window.Sends a PUT
        request to the server to update the booking details.Upon successfull update,the local booking data
        is updated,the UI is refreshed to reflect changes,and dialog window is closed.
        Args:
            booking(dict): The original booking data before update.
            engineer_dropdown(QComboBox): The combo box for selecting
                                          the Engineer.
            location_dropdown(QComboBox): The combo box for selecting
                                          the location.
            date_edit(QDateEdit): The date edit widget for selecting the date.
            duration_spinbox(QSpinBox): The spinbox for selecting duration.
            dialog(QDialog):The dialog window for updating the booking.
        Raises:
            requests.exceptions.RequestException: If an error occurs while making PUT
                                                  request to server.
        """
        new_booking_data = {
            "engineer": engineer_dropdown.currentText(),
            "location": location_dropdown.currentText(),
            "date": date_edit.date().toString("yyyy/MM/dd"),
            "duration": duration_spinbox.value()
        }
        update_api_url = "http://192.168.17.72:5000/update_booking"
        headers = {"Content-Type": "application/json"}
        payload = {
            "engineer": booking["engineer"],
            "location": booking["location"],
            "date": booking["date"],
            "duration": booking["duration"],
            "update": new_booking_data
        }
        try:
            response = requests.put(update_api_url, data=json.dumps(payload), headers=headers)
            response.raise_for_status()
            update_response = response.json()

            if "message" in update_response:
                print(update_response["message"])
                booking.update(new_booking_data)
                self.save_json("bookings", self.bookings)
                self.show_slot_info()
                dialog.accept()
            else:
                print("Failed to update booking. API response:", update_response)
                QMessageBox.warning(self, "Update Failed", "Failed to update booking. Please try again.")
        except requests.exceptions.RequestException as e:
            print(f"Error making update_booking API request: {e}")
            QMessageBox.critical(self, "Update Error", f"Error making update_booking API request: {e}")

    def show_slot_info(self, parent=None):
        """
        This method updates the inner group box layout to display information about engineers
        assigned to slots and unassigned engineers for the selected date. Retrieves booking
        information for the selected date and populates tables with assigned engineers and
        unassigned engineers. Updates the calendar colors to reflect the assigned and unassigned slots.
        Args:
            parent (QWidget): The parent QWidget. Default is None.
        Raises:
            None
        Returns:
            None
        """
        for i in reversed(range(self.inner_group_box_layout.count())):
            widgetToRemove = self.inner_group_box_layout.itemAt(i).widget()
            self.inner_group_box_layout.removeWidget(widgetToRemove)
            widgetToRemove.setParent(None)
        selected_date_str = self.calendar.selectedDate().toString('yyyy/MM/dd')
        assigned_engineers = []
        unassigned_engineers = set(self.engineers.keys())
        booking_info = None
        for booking in self.bookings:
            if selected_date_str == booking["date"]:
                booking_info = {
                    "engineer": booking["engineer"],
                    "site": booking["location"],
                    "duration": f"{booking['duration']} {'Hour' if booking['duration'] == 1 else 'Hours'}",
                    "remarks": booking.get("remarks", ""),  # Add remarks field
                    "date": booking["date"]
                }
                combination_exists = any(
                    entry and
                    entry.get('engineer') == booking_info['engineer'] and
                    entry.get('site') == booking_info['site'] and
                    entry.get('duration') == booking_info['duration'] and
                    entry.get('date') == booking_info['date']
                    for entry in assigned_engineers
                )
                if not combination_exists:
                    assigned_engineers.append(booking_info)
                    unassigned_engineers.discard(booking["engineer"])

        # Create the table widget
        assigned_table = QTableWidget(self)
        assigned_table.setColumnCount(4)  # Increase column count
        assigned_table.setHorizontalHeaderLabels(["Engineer", "Site", "Duration", "Remarks"])  # Add Remarks column
        assigned_table.setRowCount(len(assigned_engineers))

        assigned_table.setColumnWidth(0, 65)
        assigned_table.setColumnWidth(2, 60) 
        assigned_table.setColumnWidth(3, 150) 
        assigned_table.setRowCount(len(assigned_engineers))
        for row, assigned_engineer in enumerate(assigned_engineers):
            assigned_table.setItem(row, 0, QTableWidgetItem(assigned_engineer['engineer']))
            assigned_table.setItem(row, 1, QTableWidgetItem(assigned_engineer['site']))
            assigned_table.setItem(row, 2, QTableWidgetItem(assigned_engineer['duration']))
            assigned_table.setItem(row, 3, QTableWidgetItem(assigned_engineer['remarks']))  # Add Remarks item
        assigned_table.itemChanged.connect(self.save_remarks)
        self.inner_group_box_layout.addWidget(assigned_table)
        if unassigned_engineers:
            unassigned_table = QTableWidget(self)
            unassigned_table.setColumnCount(1)
            unassigned_table.setHorizontalHeaderLabels(["Unassigned Engineers"])
            unassigned_table.setRowCount(len(unassigned_engineers))
            for row, unassigned_engineer in enumerate(unassigned_engineers):
                unassigned_table.setItem(row, 0, QTableWidgetItem(unassigned_engineer))
            self.inner_group_box_layout.addWidget(unassigned_table)
            unassigned_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.update_calendar_colors()
    def save_remarks(self, item):
        """
        Save the edited remarks when the user modifies the "Remarks" column in the table.
        Args:
            item (QTableWidgetItem): The item that was edited.
        Returns:
            None
        """
        if item.column() == 3:  # Check if the edited item is in the Remarks column
            row = item.row()
            engineer_item = self.assigned_table.item(row, 0)  # Get the QTableWidgetItem for engineer
            remarks_item = self.assigned_table.item(row, 3)  # Get the QTableWidgetItem for remarks

        if engineer_item is not None and remarks_item is not None:  # Ensure items are not None
            engineer = engineer_item.text()  # Get the engineer's name
            remarks = remarks_item.text()  # Get the edited remarks


    def add_location(self):
        """
        This method opens a dialog window to input a new location name.If the user provides a name
        and confirms,the location is added to the internal data structure and saved to a JSON file.
        Additionally sends a POST request to server API to add the location.Displays success message
        if operation is successfull.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        location, ok = QInputDialog.getText(self, "Add Location", "Enter the location name:")
        if ok and location:
            self.locations[location] = []
            self.save_json("locations", self.locations)
           
            api_url = "http://192.168.17.72:5000/api/add_location"
            response = requests.post(api_url, json={"location": location})
            
            if response.status_code == 200:
                QMessageBox.information(self, "Success", "Location added successfully!")
            else:
                QMessageBox.warning(self, "Success", "Location added successfully!")

    def delete_location(self):
        """
        This method opens a dialog window to select locations from the existing locations.If the
        users select location and confirms,the location is deleted from internal data structure and
        removed from JSON file.Additionally sends a DELETE request to server API to delete the location.
        Displays a success message if the operation is successfull.
        Args:
            None
        Raises:
            None
        Returns:
            None
        """
        location, ok = QInputDialog.getItem(self, "Delete Location", "Select location to delete:", self.locations.keys(), 0, False)
        if ok and location:
            del self.locations[location]
            self.save_json("locations", self.locations)
            
            api_url = "http://192.168.17.72:5000/api/delete_location"
            response = requests.delete(api_url, json={"location": location})
            
            if response.status_code == 200:
                QMessageBox.information(self, "Success", "Location deleted successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to delete location. Please try again.")

    def delete_engineer(self):
        """
        This method opens a dialog window to select an engineer from the existing engineers.If the user selects
        an engineer and confirms,the engineer is deleted from the internal data structure and removed from the
        JSON file.Additionally removes any booked slots associated with deleted engineer and updates the slot 
        information displayed.Sends a DELETE requesr to server API to delete engineer.Displays a success message
        if operation is successfull.
        Args:
            self
        Raises:
            None
        Returns:
            None
        
        """
        engineer, ok = QInputDialog.getItem(self, "Delete Engineer", "Select engineer to delete:", self.engineers.keys(), 0, False)
        if ok and engineer:
            del self.engineers[engineer]
            self.save_json("engineers", self.engineers)
            self.booked_slots.pop(engineer, None)
            self.show_slot_info()
           
            api_url = "http://192.168.17.72:5000/api/delete_engineer"
            response = requests.delete(api_url, json={"engineer": engineer})
            
            if response.status_code == 200:
                QMessageBox.information(self, "Success", "Engineer deleted successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to delete engineer. Please try again.")

    def add_engineer(self):
        """
        This method opens a dialog window to input the name of a new engineer.If the user enters 
        a name and confirms,sends a POST request to the server API to add the engineer.Adds the
        new engineer to the internal data structure and also add to the JSON file.Displays a 
        success message if the operation is successfull.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        engineer, ok = QInputDialog.getText(self, "Add Engineer", "Enter the engineer's name:")
        if ok and engineer:

            api_url = "http://192.168.17.72:5000/api/add_engineer"
            response = requests.post(api_url, json={"engineer": engineer})
            self.engineers[engineer] = []
            self.save_json("engineers", self.engineers)
          
            
            
            if response.status_code == 200:
                QMessageBox.information(self, "Success", "Engineer added successfully!")
            else:
                QMessageBox.warning(self, "Success", "Engineer added successfully!")

    def select_location_from_dropdown(self, index):
        """
        This method retrieves the selected location from the dropdown menu based on index provided.
        Calls the select_location method with selected location as an argument.
        Args:
            index(int): The index of the selected item in the dropdown menu.
        Raises:
            None
        Returns:
            None
        """
        location = self.location_dropdown.itemText(index)
        self.select_location(location)

    def select_engineer_from_dropdown(self, index):
        """
        This method retrieves the selected engineer from the dropdown menu based on index provided.
        Calls the selected_engineer method with selected engineer as an argument.
        Args:
            index(int): The index of the selected item in the dropdown menu.
        Raises:
            None
        Returns:
            None 
        """
        engineer = self.engineer_dropdown.itemText(index)
        self.select_engineer(engineer)

    def select_location(self, location):
        """
        This method sets the current location attribute to the specified location.
        Args:
            location(str): The name of location to set.
        Raises:
            None
        Returns:
            None
        
        """
        self.current_location = location

    def select_engineer(self, engineer):
        """
        This method sets the current engineer atrribute to the specified engineer.
        Args:
            engineer(str): The name of engineer to set.
        Raises:
            None
        Returns:
            None
        """
        self.current_engineer = engineer

    def show_book_slot_dialog(self, slot={}):
        """
        This method sets the current widget to the calendar tab and diplays the time
        slot dialog.If the dialog is accepted ,it retrieves the selected data and 
        books the time slot.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        global current_slot
        dialog = TimeSlotDialog(self.locations.keys(), self.engineers.keys(), slot=slot,parent=self)
        
        if dialog.exec_():
            selected_data = dialog.get_selected_data()
            self.book_slot(selected_data)

    def book_slot(self, selected_data):
        """
        This method books a slot for an engineer at specified location and date.
        Args:
            selected_data(dict): A dictionary containing the selected slot data.
        Raises: 
            requests.exceptions.RequestException: If there is an error while making the
                                                  booking API request.
        Returns:
            None
        """
        date = self.calendar.selectedDate()
        duration = selected_data['time']
        slot = {
            'engineer': selected_data['engineer'],
            'location': selected_data['location'],
            'date': date.toString('yyyy/MM/dd'),
            'duration': duration
        }
        if any(
                booking["engineer"] == selected_data["engineer"]
                and booking["location"] == selected_data["location"]
                and booking["date"] == date.toString("yyyy/MM/dd")
                and booking["duration"] == duration
                for booking in self.bookings
        ):
            print("Duplicate entry. Booking not added.")
            QMessageBox.warning(self, "Duplicate Entry", "Booking not added. Duplicate entry.")
            return
        self.reports_widget.add_booked_data(slot)
        if selected_data['engineer'] not in self.booked_slots:
            self.booked_slots[selected_data['engineer']] = []
        if slot not in self.booked_slots[selected_data['engineer']]:
            self.booked_slots[selected_data['engineer']].append(slot)
            data = {
                "engineer": selected_data['engineer'],
                "location": selected_data['location'],
                "duration": duration,
                "date": date.toString('yyyy/MM/dd')
            }
            booking_api_url ="http://192.168.17.72:5000/api/booking" 
            headers = {"Content-Type": "application/json"}
            try:
                response = requests.post(booking_api_url, data=json.dumps(data), headers=headers)
                response.raise_for_status()
                booking_response = response.json()

                if "booking_id" in booking_response:
                    booking_id = booking_response["booking_id"]
                    print("Booking successful! Booking ID:", booking_id)
                    data["booking_id"] = booking_id
                    QMessageBox.information(self, "Booking Successful", f"Booking successful! Booking ID: {booking_id}")
                else:
                    print("Booking failed. API response:", booking_response)
                    QMessageBox.warning(self, "Booking Failed", "Booking failed. Please try again.")
            except requests.exceptions.RequestException as e:
                print(f"Error making booking API request: {e}")
                QMessageBox.critical(self, "Booking Error", f"Error making booking API request: {e}")
            self.bookings.append(data)
            self.update_location_label(selected_data['location'])
            self.save_json("bookings", self.bookings)
        else:
            print("Duplicate entry.Booking not added.")
            QMessageBox.warning(self, "Duplicate Entry", "Booking not added. Duplicate entry.")
        date = self.calendar.selectedDate()
        duration = selected_data['time']
        slot = {
                'engineer': selected_data['engineer'],
                'location': selected_data['location'],
                'date': date.toString('yyyy/MM/dd'),
                'duration': duration,
                'Created Date': dt.datetime.now().strftime('%Y/%m/%d')  
        }
        self.reports_widget.add_booked_data(slot)     
    
    def update_location_label(self, location):
        """
        This method updates the label with information about booked slots for a given
        location.
        Args:
            location(str): The name of the location.
        Raises:
            None
        Returns:
            None
        """
        if location in self.locations:
            self.booked_slots = {}
            booked_slots = "\n".join(self.locations[location])
        # self.location_label.setText(f"Location: {location}\nBooked Slots:\n{booked_slots}")
        else:
        # Handle the case when the location is not found
            booked_slots = "No bookings for this location"

    def update_dashboard(self):
        """
        This method updates the dashboard with latest booking data.
        Args:
            Self
        Raises:
            None
        Returns:
            None
        """
        self.dashboard_tab.update_charts(self.bookings)
    
    def save_json(self, what_to_save, data):
        """
        This method saves the provided data as a JSON.
        Args:
            what_to_save(str):Indicates which type of data to save.
            data(dict): The data to be saved.
        Raises:
            None
        Returns:
            None
        """
        # save json
        self.data[what_to_save] = data

        with open ("booked_slot.json","w+") as json_file:
            json.dump(self.data,json_file,indent=4)
    
    def initialize_calendar_colors(self):
        """
        This function iterates through the existing booking and calculatr the total durations for
        assigned and unassigned slots for each date.it then updates the calendar accordingly
        with fully assigned slots displayed in red and unassigned slots in green.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        self.calendar_date_formats = {}
        for booking in self.bookings:                  
            date = booking["date"]
            if date not in self.calendar_date_formats:
                self.calendar_date_formats[date] = {"assigned": 0, "unassigned": 0}

            if booking["engineer"] in self.booked_slots:
                # Fully assigned slots in red
                self.calendar_date_formats[date]["assigned"] += booking["duration"]
            else:
                # Unassigned slots in green
                self.calendar_date_formats[date]["unassigned"] += booking["duration"]

        selected_date = self.calendar.selectedDate().toString('yyyy/MM/dd')
        if selected_date not in self.calendar_date_formats:
            self.calendar_date_formats[selected_date] = {"assigned": 0, "unassigned": 0}
        self.update_calendar_colors()

    def fetch_data_from_api(self, from_date=None, to_date=None):
        """
        This method sends POST request to the specified API URL with optional date range parameters
        to fetch data.
        Args:
            from_date(str): The start date of the range in 'yyyy/MM/dd' format.Default is None.
            to_date(str): The end date of the range in 'yyyy/MM/dd' format.Default is None.
        Raises:
            Exception: If an error occurs while fetching data from API,such as network issues
                       or invalid response.
        Returns: 
            dict: A dictionary containing the fetched data from API.
        """
        api_url = "http://192.168.17.72:5000/api/fetchdata"
        payload = {"from": from_date, "to": to_date}
        headers = {"Content-Type": "application/json"}
        try:
            response = requests.post(api_url, json=payload, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data from API: {e}")
            return {"status": "Failure", "error_code": 500, "message": "Error fetching data from API", "data": {}}

    def load_json(self):
        """
        This method attempts to fetch data from API.If successfull retrieves bookings,engineers and location
        from API response.If the 'locations' and 'engineer' keys are missing from the API response,loads data
        from local JSON file named 'booked_slot.json'.
        Args:
            self
        Raises:
            Exception: If an error occurs while loading data from API or local file.
        Returns:
            dict: A dictionary containing bookings,engineers and locations.
        """
        try:
            data = self.fetch_data_from_api()["data"]
            bookings = data.get("bookings", [])
            
            if "locations" in data and "engineers" in data:
                locations = data["locations"]
                engineers = data["engineers"]
            else:
                print("No 'locations' and 'engineers' keys found in API response. Loading from local file.")
                with open("booked_slot.json") as json_file:
                    local_data = json.load(json_file)

                locations = local_data.get("locations", [])
                engineers = local_data.get("engineers", [])

            print({"bookings": bookings, "engineers": engineers, "locations": locations})
            return {"bookings": bookings, "engineers": engineers, "locations": locations}

        except Exception as e:
            print(f"An exception occurred while loading data: {e}")
            return {"bookings": [], "engineers": [], "locations": []}
        
    def showEvent(self, event):
        """
        This method overrides the QWidget showEvent method to initialize calendar colors.
        Args:
            event: The event parameter.
        Raises:
            None
        Returns:
            None
        """
        super().showEvent(event)
        self.initialize_calendar_colors()

    def refresh_fetchdata(self):
        """
        Fetches the latest data from the API and updates the local data.
        Updates the slot information and calendar colors accordingly.
        Args:
            Self
        Raises:
            None
        Returns:
            None
        """
        api_data = self.fetch_data_from_api()
        self.data["bookings"] = self.data.get("data", [])
        self.show_slot_info()
        self.update_calendar_colors()
        if api_data["status"] == "Success":
            self.log_message("Data has been refreshed successfully.")
        else:
            self.log_message("Failed to refresh data. Please try again.")

    def log_message(self, message):
       """
       This method Logs a message to the console.
       Args:
            message (str): The message to be logged.
        Raises:
            None
        Returns:
            None
       """
       print(message) 

    def start_timer(self):
        """
        The timer emits a timeout signal at regular intervals, triggering the complaint_timer method.
        Args:
            self
        Raises:
            None
        Returns:
            None

        """
        self.timer = QTimer(self)
        complaint_timer = 1000
        self.timer.timeout.connect(self.complaint_timer)
        self.timer.start(complaint_timer)

    def complaint_timer(self):
        """
        This method is called when the complaint timer times out. It triggers the display of the book slot dialog 
        if the slot was booked from a complaint.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        global current_slot
        global book_from_complaint
        if book_from_complaint==True:
            book_from_complaint=False
            self.show_book_slot_dialog(slot=current_slot)

# book_from_complaint = False        
class LoginWindow(QDialog):
    """
    This class represents a login window dialog.
    """
    login_signal = pyqtSignal()
    def __init__(self, parent=None):
        """
        This method initializes the Login Window.
        Args:
            parent(QWidget); The parents widget of the dialog.Default is None.
        Raises:
            None
        Returns:
            None
        """
        super().__init__(parent)
        self.setWindowTitle("Login")
        self.setGeometry(500, 200, 300, 200)
        layout = QVBoxLayout(self)
        self.email_label = QLabel("Email:")
        self.email_input = QLineEdit(self)
        self.password_label = QLabel("Password:")
        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.Password)
        self.login_button = QPushButton("Login", self)
        self.login_button.clicked.connect(self.login)
        layout.addWidget(self.email_label)
        layout.addWidget(self.email_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.login_button)
        self.setLayout(layout)

    def login(self):
        """
        This method checks the entered email and password against predefined credentials.
        If they match,emits the login signal and close the login window.Otherwise,display
        a warning message.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        email = self.email_input.text()
        password = self.password_input.text()
        if email == "user" and password == "123":
            self.login_signal.emit()  # Emit the login signal
            self.close()
        else:
            error_message = "Invalid credentials. Please try again."
            QMessageBox.warning(self, "Login Error", error_message)

    def login(self):
        """
        This method sends a login request to the API with the entered Email and Password.
        If the response status code is 200 and the message indicates a successfull login,
        emits the login signal,closes the login window and creates a license file.Otherwise
        display a warning message.
        Args:
            self
        Raises:
            None
        Returns:
            None
        """
        email = self.email_input.text()
        password = self.password_input.text()
        # Make the login API request
        login_url = "http://192.168.17.72:5000/api/login"
        payload = {"email": email, "password": password}
        headers = {"Content-Type": "application/json"}
        response = requests.post(login_url, data=json.dumps(payload), headers=headers)
        print("Response Status Code:", response.status_code)
        print("Response Content:", response.content)
        
        if response.status_code == 200 and response.json().get("message") == "Login successful":
            self.login_signal.emit()
            self.accept()
            self.create_license_file(email)
        else:
            error_message = "Invalid credentials. Please try again."
            QMessageBox.warning(self, "Login Error", error_message)

    def load_existing_credentials(self):
        """
        This method attempts to open and read 'license.json' file.If the file is found
        it loads the existing data and checks if any user has an "active" status.If found
        returns True,otherwise returns False.
        Args:
            self
        Raises:
            FileNotFoundError: If 'license.json' file not found.
        Returns:
            bool: True if active user credentials are found,False otherwise
        """
        try:
            with open("license.json", "r") as license_file:
                existing_data = json.load(license_file)
                for i in existing_data["user"]:
                    if i["status"] == "active":
                        return True
                return False
        except FileNotFoundError:
            return False

    def create_license_file(self, email, is_active=True):
        """
        This method constructs a license entry containing user information such as user name,
        email,datetime and status.If the file is found,it loads the existing data and updates 
        the user entry if it's already exists,otherwise it appends the new user entry.If the 
        file is not found,it creates the new file and writes the user entry to it.
        Args:
            email(str): The email address of the user.
            is_active(bool): Indicates whether the user is active.
        Raises:
            FileNotFoundError: If the 'license.json' file not found.
        Returns:
            None
        """
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = email.split('@')[0][:6]
        license_entry = {
            "user_name": user_name,
            "email": email,
            "datetime": current_datetime,
            "status": "active" if is_active else "inactive"
        }
        try:
            with open("license.json", "r") as license_file:
                existing_data = json.load(license_file)
            user_entry = next((user for user in existing_data.get("user", []) if user["email"] == email), None)
            if user_entry:
                user_entry.update(license_entry)
            else:
                existing_data.setdefault("user", []).append(license_entry)
            with open("license.json", "w") as license_file:
                json.dump(existing_data, license_file, indent=4)
        except FileNotFoundError:
            with open("license.json", "w") as license_file:
                json.dump({"user": [license_entry]}, license_file, indent=4)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_app = AppointmentSchedulerApp()
    sys.exit(app.exec_())

